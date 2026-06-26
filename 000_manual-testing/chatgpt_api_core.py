"""
Core ChatGPT Responses API logic: client creation, sendMessage, logging, cost estimation.
Used by the CLI (chatgpt-API-call-tester.py) and by experiment runners (run_chatgpt_experiments.py).
"""
import os
import sys
import time
from datetime import datetime

_scriptDir = os.path.dirname(os.path.abspath(__file__))
ENV_LOCAL_PATH = os.path.join(_scriptDir, "..", ".env.local")
API_KEY_ENV_VAR = "NEXT_PUBLIC_OPENAI_API_KEY"
DEFAULT_MODEL = "gpt-5.2"
LOG_HEADERS = ("Name", "Timestamp", "Input", "Output", "Input tokens", "Output tokens", "Est. cost ($)", "Time (s)")
LOG_FILENAME = "chatgpt-API-call-tester_log.xlsx"

MODEL_PRICING = {
    "gpt-4o": (2.50, 10.00),
    "gpt-4o-mini": (0.15, 0.60),
    "gpt-4-turbo": (10.00, 30.00),
    "gpt-4": (30.00, 60.00),
    "gpt-3.5-turbo": (0.50, 1.50),
    "gpt-5.2": (1.75, 14.00),
    "gpt-5.4": (2.50, 15.00),
}


def loadEnvLocal():
    """Load KEY=VALUE from ../.env.local into os.environ."""
    if not os.path.isfile(ENV_LOCAL_PATH):
        return
    with open(ENV_LOCAL_PATH) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, value = line.partition("=")
                os.environ[key.strip()] = value.strip().strip("'\"")


def getLogPath():
    """Path for the Excel log file (constant filename)."""
    return os.path.join(_scriptDir, LOG_FILENAME)


def estimateCost(inputTokens, outputTokens, model):
    """Estimate cost in USD from token counts. Returns float or None if model unknown."""
    if (inputTokens or 0) == 0 and (outputTokens or 0) == 0:
        return 0.0
    baseModel = model.split("/")[-1].split("-202")[0] if model else ""
    prices = MODEL_PRICING.get(baseModel) or MODEL_PRICING.get(model)
    if not prices:
        return None
    inputPerM, outputPerM = prices
    return (inputTokens or 0) * inputPerM / 1e6 + (outputTokens or 0) * outputPerM / 1e6


def appendRowToExcel(logPath, row):
    """Append a row to the Excel log file; create file with headers if missing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl for logging: pip install openpyxl", file=sys.stderr)
        return
    if os.path.isfile(logPath):
        wb = load_workbook(logPath)
        ws = wb.active
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chat log"
        ws.append(LOG_HEADERS)
    ws.append(row)
    wb.save(logPath)


def createClient(apiKey=None):
    """Create OpenAI client. Uses env if apiKey not provided."""
    loadEnvLocal()
    key = apiKey or os.environ.get(API_KEY_ENV_VAR) or os.environ.get("OPENAI_API_KEY")
    if not key:
        raise ValueError(f"Set {API_KEY_ENV_VAR} in .env.local or pass api_key")
    from openai import OpenAI
    return OpenAI(api_key=key)


def extractOutputText(response):
    """Fallback: aggregate text from response.output message items."""
    out = []
    for item in getattr(response, "output", []) or []:
        if getattr(item, "type", None) != "message":
            continue
        for content in getattr(item, "content", []) or []:
            if getattr(content, "type", None) == "output_text":
                out.append(getattr(content, "text", "") or "")
    return "".join(out).strip() or None


def sendMessage(
    client,
    inputText,
    *,
    model=DEFAULT_MODEL,
    reasoning="none",
    temperature=None,
    codeInterpreter=False,
    codeInterpreterMemory="4g",
    conversationId=None,
):
    """
    Send one message to the Responses API and return (outputText, usage, elapsedSec, cost).
    reasoning: "none" | "high" (for "high", temperature is omitted per API guidance).
    """
    kwargs = {
        "model": model,
        "input": inputText,
        "reasoning": {"effort": reasoning},
    }
    if temperature is not None:
        kwargs["temperature"] = temperature
    if codeInterpreter:
        kwargs["tools"] = [
            {"type": "code_interpreter", "container": {"type": "auto", "memory_limit": codeInterpreterMemory}},
        ]
    if conversationId is not None:
        kwargs["conversation"] = {"id": conversationId}

    startTime = time.perf_counter()
    response = client.responses.create(**kwargs)
    elapsed = time.perf_counter() - startTime

    text = getattr(response, "output_text", None) or extractOutputText(response)
    usage = getattr(response, "usage", None)
    inputTokens = getattr(usage, "input_tokens", None) if usage else None
    outputTokens = getattr(usage, "output_tokens", None) if usage else None
    cost = estimateCost(inputTokens, outputTokens, model) if usage else None

    return {
        "outputText": text or "",
        "inputTokens": inputTokens,
        "outputTokens": outputTokens,
        "elapsedSec": elapsed,
        "cost": cost,
        "response": response,
    }
