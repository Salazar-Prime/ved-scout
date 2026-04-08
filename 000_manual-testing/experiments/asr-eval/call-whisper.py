import os
import time
import requests
import openpyxl
from pathlib import Path

BASE_URL = "http://localhost:3000/api/transcribe"
DATA_DIR = Path(__file__).parent / "data"
GROUND_TRUTH_FILE = DATA_DIR / "ASR-Ground-truth.xlsx"
OUTPUT_FILE = Path(__file__).parent / "output" / "asr-eval-log.xlsx"


def loadGroundTruth():
    wb = openpyxl.load_workbook(GROUND_TRUTH_FILE)
    ws = wb.active
    groundTruth = {}
    for row in ws.iter_rows(values_only=True):
        expectedText, audioFilename = row
        groundTruth[audioFilename] = expectedText
    return groundTruth


def transcribeFile(audioPath: Path):
    with open(audioPath, "rb") as f:
        files = {"audio": (audioPath.name, f, "audio/mpeg")}
        startTime = time.time()
        response = requests.post(BASE_URL, files=files)
        elapsed = round(time.time() - startTime, 2)
    return response, elapsed


def main():
    groundTruth = loadGroundTruth()

    audioFiles = sorted(DATA_DIR.glob("*.mp3"))
    print(f"Found {len(audioFiles)} audio files")
    print(f"Ground truth entries: {len(groundTruth)}")
    print("-" * 60)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASR Eval Log"
    headers = [
        "File", "Ground Truth", "Whisper Response",
        "Status", "Latency (s)", "Audio Duration (s)", "Language", "Segments"
    ]
    ws.append(headers)

    for audioFile in audioFiles:
        filename = audioFile.name
        expected = groundTruth.get(filename, "N/A")
        print(f"Transcribing: {filename} ... ", end="", flush=True)

        audioDuration = ""
        language = ""
        numSegments = ""

        try:
            response, elapsed = transcribeFile(audioFile)
            if response.status_code == 200:
                data = response.json()
                transcribedText = data.get("text", "")
                audioDuration = data.get("durationInSeconds", "")
                language = data.get("language", "")
                segments = data.get("segments", [])
                numSegments = len(segments) if segments else 0
                status = "OK"
            else:
                transcribedText = f"Error {response.status_code}: {response.text}"
                status = "ERROR"
        except Exception as e:
            transcribedText = str(e)
            status = "EXCEPTION"
            elapsed = 0

        print(f"{status} ({elapsed}s)")
        ws.append([
            filename, expected, transcribedText,
            status, elapsed, audioDuration, language, numSegments
        ])

    wb.save(OUTPUT_FILE)
    print("-" * 60)
    print(f"Log saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
