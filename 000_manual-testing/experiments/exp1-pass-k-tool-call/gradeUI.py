#!/usr/bin/env python3
"""
Simple Tkinter grading UI for reviewing Pass^k experiment results.
Loads "Question Set" (expected) and "Summary" (actual) from an Excel workbook,
lets you compare side-by-side, and mark each run as pass / fail / for varun / revisit.
Grades persist to a "Grade" column on the Summary sheet.
"""
import json
import os
import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, "output", "Exp-Pass-k-calc copy.xlsx")

GRADE_OPTIONS = ("pass", "fail", "for varun", "revisit")

_RUN_NUM_RE = re.compile(r"-run(\d+)-")

GRADE_COLORS = {
    "pass": "#22883e",
    "fail": "#cc2222",
    "for varun": "#cc9a00",
    "revisit": "#cc9a00",
    "": "#3366cc",
}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _normalizeHeaders(row):
    """Return list of stripped-lowercase header strings."""
    return [str(c).strip().lower() if c else "" for c in row]


def loadQuestionSet(ws) -> Dict[str, Dict[str, Any]]:
    """Load Question Set sheet into a dict keyed by question ID."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = _normalizeHeaders(rows[0])
    questions: Dict[str, Dict[str, Any]] = {}
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and headers[i]}
        qId = str(d.get("id", "")).strip()
        if qId:
            questions[qId] = d
    return questions


def loadSummary(ws) -> tuple:
    """Load Summary sheet. Returns (headersList, list-of-row-dicts, gradeColIndex)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], None
    rawHeaders = [str(c).strip() if c else "" for c in rows[0]]
    normHeaders = [h.lower() for h in rawHeaders]

    gradeColIdx = None
    if "grade" in normHeaders:
        gradeColIdx = normHeaders.index("grade")

    summaryRows: List[Dict[str, Any]] = []
    for rowIdx, row in enumerate(rows[1:], start=1):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d: Dict[str, Any] = {"_excelRow": rowIdx + 1}  # 1-indexed excel row (header=1)
        for i in range(len(row)):
            if i < len(rawHeaders) and rawHeaders[i]:
                d[rawHeaders[i]] = row[i]
        if gradeColIdx is not None and gradeColIdx < len(row):
            d["_grade"] = str(row[gradeColIdx]).strip() if row[gradeColIdx] else ""
        else:
            d["_grade"] = ""
        summaryRows.append(d)
    return rawHeaders, summaryRows, gradeColIdx


# ---------------------------------------------------------------------------
# Grade persistence
# ---------------------------------------------------------------------------

def saveGrade(filePath: str, excelRow: int, grade: str, rawHeaders: List[str]):
    """Write grade into the Summary sheet and save the workbook."""
    wb = load_workbook(filePath)
    ws = wb["Summary"]

    headerCells = [str(c.value).strip() if c.value else "" for c in ws[1]]
    normHeaders = [h.lower() for h in headerCells]

    if "grade" in normHeaders:
        colIdx = normHeaders.index("grade") + 1  # openpyxl is 1-indexed
    else:
        colIdx = len(headerCells) + 1
        ws.cell(row=1, column=colIdx, value="Grade")

    ws.cell(row=excelRow, column=colIdx, value=grade)
    wb.save(filePath)
    wb.close()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

class GradeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Pass^k Grading UI")
        self.geometry("1300x800")
        self.minsize(900, 550)

        self.filePath: Optional[str] = None
        self.questions: Dict[str, Dict[str, Any]] = {}
        self.summaryHeaders: List[str] = []
        self.runs: List[Dict[str, Any]] = []
        self.gradeColIdx: Optional[int] = None
        self.currentIdx: int = -1

        self._buildWidgets()

        if os.path.isfile(DEFAULT_XLSX):
            self.filePath = DEFAULT_XLSX
            self.fileVar.set(DEFAULT_XLSX)
            self._loadFile()

    # ----- widget construction -----

    def _buildWidgets(self):
        # top bar: file selector
        topFrame = ttk.Frame(self, padding=4)
        topFrame.pack(fill=tk.X)

        ttk.Label(topFrame, text="Excel:").pack(side=tk.LEFT)
        self.fileVar = tk.StringVar(value=DEFAULT_XLSX)
        ttk.Entry(topFrame, textvariable=self.fileVar, width=70).pack(side=tk.LEFT, padx=4)
        ttk.Button(topFrame, text="Open...", command=self._browseFile).pack(side=tk.LEFT)
        ttk.Button(topFrame, text="Reload", command=self._loadFile).pack(side=tk.LEFT, padx=4)

        # main paned: left list | right detail
        paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # left: run list
        leftFrame = ttk.Frame(paned, width=280)
        paned.add(leftFrame, weight=0)

        ttk.Label(leftFrame, text="Runs", font=("TkDefaultFont", 11, "bold")).pack(anchor=tk.W)

        filterFrame = ttk.Frame(leftFrame)
        filterFrame.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(filterFrame, text="Filter:").pack(side=tk.LEFT)
        self.filterVar = tk.StringVar()
        filterCombo = ttk.Combobox(filterFrame, textvariable=self.filterVar, state="readonly", width=18)
        filterCombo.pack(side=tk.LEFT, padx=4)
        filterCombo.bind("<<ComboboxSelected>>", lambda _: self._populateRunList())
        self.filterCombo = filterCombo

        listFrame = ttk.Frame(leftFrame)
        listFrame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(listFrame, orient=tk.VERTICAL)
        self.runListbox = tk.Listbox(listFrame, yscrollcommand=scrollbar.set, font=("TkFixedFont", 10))
        scrollbar.config(command=self.runListbox.yview)
        self.runListbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.runListbox.bind("<<ListboxSelect>>", self._onRunSelect)

        # right: detail
        rightFrame = ttk.Frame(paned)
        paned.add(rightFrame, weight=1)

        # detail is split: expected (top) | actual (bottom)
        detailPaned = ttk.PanedWindow(rightFrame, orient=tk.VERTICAL)
        detailPaned.pack(fill=tk.BOTH, expand=True)

        bigFont = ("TkDefaultFont", 18, "bold")
        medFont = ("TkDefaultFont", 13)

        # ---------- expected pane ----------
        expFrame = ttk.LabelFrame(rightFrame, text="Expected (Question Set)", padding=4)
        detailPaned.add(expFrame, weight=1)

        self.expectedText = tk.Text(expFrame, wrap=tk.WORD, state=tk.DISABLED, height=10, font=("TkFixedFont", 10))
        expScroll = ttk.Scrollbar(expFrame, orient=tk.VERTICAL, command=self.expectedText.yview)
        self.expectedText.configure(yscrollcommand=expScroll.set)
        self.expectedText.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        expScroll.pack(side=tk.LEFT, fill=tk.Y)

        expStatsFrame = ttk.Frame(expFrame, padding=8, width=180)
        expStatsFrame.pack(side=tk.RIGHT, fill=tk.Y)
        expStatsFrame.pack_propagate(False)

        ttk.Label(expStatsFrame, text="Question ID", font=medFont).pack(anchor=tk.W)
        self.expQidLabel = ttk.Label(expStatsFrame, text="—", font=bigFont)
        self.expQidLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(expStatsFrame, text="Total Steps", font=medFont).pack(anchor=tk.W)
        self.expStepsLabel = ttk.Label(expStatsFrame, text="—", font=bigFont)
        self.expStepsLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(expStatsFrame, text="Unique Tool Types", font=medFont).pack(anchor=tk.W)
        self.expToolTypesLabel = ttk.Label(expStatsFrame, text="—", font=bigFont)
        self.expToolTypesLabel.pack(anchor=tk.W)

        # ---------- actual pane ----------
        actFrame = ttk.LabelFrame(rightFrame, text="Actual (Summary)", padding=4)
        detailPaned.add(actFrame, weight=2)

        self.actualText = tk.Text(actFrame, wrap=tk.WORD, state=tk.DISABLED, height=18, font=("TkFixedFont", 10))
        actScroll = ttk.Scrollbar(actFrame, orient=tk.VERTICAL, command=self.actualText.yview)
        self.actualText.configure(yscrollcommand=actScroll.set)
        self.actualText.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        actScroll.pack(side=tk.LEFT, fill=tk.Y)

        actStatsFrame = ttk.Frame(actFrame, padding=8, width=180)
        actStatsFrame.pack(side=tk.RIGHT, fill=tk.Y)
        actStatsFrame.pack_propagate(False)

        ttk.Label(actStatsFrame, text="Model", font=medFont).pack(anchor=tk.W)
        self.actModelLabel = ttk.Label(actStatsFrame, text="—", font=bigFont)
        self.actModelLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(actStatsFrame, text="Reasoning", font=medFont).pack(anchor=tk.W)
        self.actReasoningLabel = ttk.Label(actStatsFrame, text="—", font=bigFont)
        self.actReasoningLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(actStatsFrame, text="Run #", font=medFont).pack(anchor=tk.W)
        self.actRunNumLabel = ttk.Label(actStatsFrame, text="—", font=bigFont)
        self.actRunNumLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(actStatsFrame, text="Question ID", font=medFont).pack(anchor=tk.W)
        self.actQidLabel = ttk.Label(actStatsFrame, text="—", font=bigFont)
        self.actQidLabel.pack(anchor=tk.W, pady=(0, 12))

        ttk.Label(actStatsFrame, text="Total Tool Calls", font=medFont).pack(anchor=tk.W)
        self.actToolCallsLabel = ttk.Label(actStatsFrame, text="—", font=bigFont)
        self.actToolCallsLabel.pack(anchor=tk.W)

        # bottom bar: grade buttons + nav + status
        bottomFrame = ttk.Frame(self, padding=4)
        bottomFrame.pack(fill=tk.X)

        ttk.Button(bottomFrame, text="< Prev", command=self._prevRun).pack(side=tk.LEFT)
        ttk.Button(bottomFrame, text="Next >", command=self._nextRun).pack(side=tk.LEFT, padx=(4, 16))

        ttk.Button(bottomFrame, text="Pass (1)", command=lambda: self._setGrade("pass")).pack(side=tk.LEFT, padx=2)
        ttk.Button(bottomFrame, text="Fail (2)", command=lambda: self._setGrade("fail")).pack(side=tk.LEFT, padx=2)
        ttk.Button(bottomFrame, text="For Varun (3)", command=lambda: self._setGrade("for varun")).pack(side=tk.LEFT, padx=2)
        ttk.Button(bottomFrame, text="Revisit (4)", command=lambda: self._setGrade("revisit")).pack(side=tk.LEFT, padx=2)

        self.statusVar = tk.StringVar(value="No file loaded")
        ttk.Label(bottomFrame, textvariable=self.statusVar).pack(side=tk.RIGHT)

        # keyboard shortcuts
        self.bind("1", lambda _: self._setGrade("pass"))
        self.bind("2", lambda _: self._setGrade("fail"))
        self.bind("3", lambda _: self._setGrade("for varun"))
        self.bind("4", lambda _: self._setGrade("revisit"))
        self.bind("n", lambda _: self._nextRun())
        self.bind("p", lambda _: self._prevRun())

    # ----- file ops -----

    def _browseFile(self):
        path = filedialog.askopenfilename(
            initialdir=os.path.join(SCRIPT_DIR, "output"),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.fileVar.set(path)
            self.filePath = path
            self._loadFile()

    def _loadFile(self):
        path = self.fileVar.get()
        if not path or not os.path.isfile(path):
            messagebox.showerror("Error", f"File not found:\n{path}")
            return
        self.filePath = path

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open workbook:\n{e}")
            return

        if "Question Set" not in wb.sheetnames:
            messagebox.showerror("Error", "'Question Set' sheet not found.")
            wb.close()
            return
        if "Summary" not in wb.sheetnames:
            messagebox.showerror("Error", "'Summary' sheet not found.")
            wb.close()
            return

        self.questions = loadQuestionSet(wb["Question Set"])
        self.summaryHeaders, self.runs, self.gradeColIdx = loadSummary(wb["Summary"])
        wb.close()

        gradeFilterValues = ["All", "Ungraded", "pass", "fail", "for varun", "revisit"]
        self.filterCombo["values"] = gradeFilterValues
        self.filterVar.set("All")

        self._populateRunList()
        self.statusVar.set(f"Loaded {len(self.runs)} runs, {len(self.questions)} questions  |  {os.path.basename(path)}")

    # ----- run list -----

    def _filteredIndices(self) -> List[int]:
        """Return indices into self.runs matching the current filter."""
        filt = self.filterVar.get()
        indices = []
        for i, r in enumerate(self.runs):
            grade = r.get("_grade", "")
            if filt == "All":
                indices.append(i)
            elif filt == "Ungraded" and grade == "":
                indices.append(i)
            elif grade == filt:
                indices.append(i)
        return indices

    def _populateRunList(self):
        self.runListbox.delete(0, tk.END)
        self._visibleIndices = self._filteredIndices()
        for listIdx, vi in enumerate(self._visibleIndices):
            r = self.runs[vi]
            qId = r.get("Question ID", "?")
            runId = r.get("Run ID", "")
            grade = r.get("_grade", "")
            tag = f" [{grade}]" if grade else ""
            label = f"#{vi+1}  {qId}  {runId}{tag}"
            self.runListbox.insert(tk.END, label)
            self.runListbox.itemconfig(listIdx, fg=GRADE_COLORS.get(grade, GRADE_COLORS[""]))

        if self._visibleIndices:
            self.runListbox.selection_set(0)
            self.runListbox.event_generate("<<ListboxSelect>>")
        else:
            self.currentIdx = -1
            self._clearDetail()

    # ----- run selection -----

    def _onRunSelect(self, _event=None):
        sel = self.runListbox.curselection()
        if not sel:
            return
        listIdx = sel[0]
        if listIdx >= len(self._visibleIndices):
            return
        self.currentIdx = self._visibleIndices[listIdx]
        self._showDetail(self.currentIdx)

    def _prevRun(self):
        sel = self.runListbox.curselection()
        if not sel or sel[0] == 0:
            return
        newSel = sel[0] - 1
        self.runListbox.selection_clear(0, tk.END)
        self.runListbox.selection_set(newSel)
        self.runListbox.see(newSel)
        self.runListbox.event_generate("<<ListboxSelect>>")

    def _nextRun(self):
        sel = self.runListbox.curselection()
        if not sel or sel[0] >= self.runListbox.size() - 1:
            return
        newSel = sel[0] + 1
        self.runListbox.selection_clear(0, tk.END)
        self.runListbox.selection_set(newSel)
        self.runListbox.see(newSel)
        self.runListbox.event_generate("<<ListboxSelect>>")

    # ----- detail display -----

    def _clearDetail(self):
        for w in (self.expectedText, self.actualText):
            w.configure(state=tk.NORMAL)
            w.delete("1.0", tk.END)
            w.configure(state=tk.DISABLED)
        for lbl in (self.expQidLabel, self.expStepsLabel, self.expToolTypesLabel,
                     self.actModelLabel, self.actReasoningLabel, self.actRunNumLabel, self.actQidLabel, self.actToolCallsLabel):
            lbl.configure(text="—")

    def _showDetail(self, idx: int):
        run = self.runs[idx]
        qId = str(run.get("Question ID", "")).strip()
        question = self.questions.get(qId)

        # expected — sidebar
        self.expQidLabel.configure(text=qId or "—")
        if question:
            self.expStepsLabel.configure(text=str(question.get("total steps", "—")))
            self.expToolTypesLabel.configure(text=str(question.get("unique tool types", "—")))
        else:
            self.expStepsLabel.configure(text="—")
            self.expToolTypesLabel.configure(text="—")

        # expected — text
        self.expectedText.configure(state=tk.NORMAL)
        self.expectedText.delete("1.0", tk.END)
        if question:
            lines = [
                f"User Prompt:           {question.get('user prompt', '')}",
                f"Follow-up 1:           {question.get('follow-up 1', '')}",
                f"Follow-up 2:           {question.get('follow-up 2', '')}",
                "",
                f"Expected Call Sequence:",
                f"  {question.get('expected call sequence', '')}",
            ]
            self.expectedText.insert(tk.END, "\n".join(lines))
        else:
            self.expectedText.insert(tk.END, f"(No question found for ID: {qId})")
        self.expectedText.configure(state=tk.DISABLED)

        # actual — sidebar
        self.actModelLabel.configure(text=str(run.get("Model", "—")))
        self.actReasoningLabel.configure(text=str(run.get("Reasoning", "—")))
        runId = str(run.get("Run ID", ""))
        m = _RUN_NUM_RE.search(runId)
        runNum = m.group(1) if m else "—"
        self.actRunNumLabel.configure(text=runNum)
        self.actQidLabel.configure(text=qId or "—")
        self.actToolCallsLabel.configure(text=str(run.get("Total Tool Calls", "—")))

        # actual — text
        self.actualText.configure(state=tk.NORMAL)
        self.actualText.delete("1.0", tk.END)

        grade = run.get("_grade", "")
        lines = [
            f"Run ID:                {run.get('Run ID', '')}",
            f"Reasoning:             {run.get('Reasoning', '')}",
            f"Total User Messages:   {run.get('Total User Messages', '')}",
            f"Current Grade:         {grade if grade else '(none)'}",
            "",
            f"Interaction Sequence:",
            f"  {run.get('Interaction Sequence', '')}",
            "",
            "Turn Details (JSON):",
        ]
        self.actualText.insert(tk.END, "\n".join(lines))

        turnDetailsRaw = run.get("Turn Details (JSON)", "")
        if turnDetailsRaw:
            try:
                parsed = json.loads(str(turnDetailsRaw))
                pretty = json.dumps(parsed, indent=2, ensure_ascii=False)
            except (json.JSONDecodeError, TypeError):
                pretty = str(turnDetailsRaw)
            self.actualText.insert(tk.END, "\n" + pretty)
        else:
            self.actualText.insert(tk.END, "\n(empty)")

        self.actualText.configure(state=tk.DISABLED)

    # ----- grading -----

    def _setGrade(self, grade: str):
        if self.currentIdx < 0 or not self.filePath:
            return
        run = self.runs[self.currentIdx]
        excelRow = run["_excelRow"]

        try:
            saveGrade(self.filePath, excelRow, grade, self.summaryHeaders)
        except PermissionError:
            messagebox.showerror("Error", "Cannot save — file may be open in another application. Close it and retry.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Save failed:\n{e}")
            return

        run["_grade"] = grade

        # refresh list label and color for current item
        sel = self.runListbox.curselection()
        if sel:
            listIdx = sel[0]
            qId = run.get("Question ID", "?")
            runId = run.get("Run ID", "")
            tag = f" [{grade}]" if grade else ""
            label = f"#{self.currentIdx+1}  {qId}  {runId}{tag}"
            self.runListbox.delete(listIdx)
            self.runListbox.insert(listIdx, label)
            self.runListbox.itemconfig(listIdx, fg=GRADE_COLORS.get(grade, GRADE_COLORS[""]))
            self.runListbox.selection_set(listIdx)

        self._showDetail(self.currentIdx)
        self.statusVar.set(f"Marked run #{self.currentIdx+1} as '{grade}'  — saved")

        self.after(150, self._nextRun)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = GradeApp()
    app.mainloop()
