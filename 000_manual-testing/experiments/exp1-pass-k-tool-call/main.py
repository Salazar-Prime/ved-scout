#!/usr/bin/env python3
"""
Pass^k Tool-Call Experiment Runner
Calls flight-assistant API (Option A), logs user messages and tool calls per run.
Supports multiple models and reasoning configurations.
"""
import argparse
import json
import os
import random
import string
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

# HTTP client
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    import urllib.request
    import urllib.parse
    import urllib.error
    HAS_REQUESTS = False

# Excel logging
try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)


# ============================================================================
# Model Configuration
# ============================================================================

DEFAULT_MODEL = "gpt-5.2"


# ============================================================================
# Cache Busting
# ============================================================================

def addCacheBuster(message: str) -> str:
    """
    Prepend a random cache-busting string to avoid API response caching.
    Format: [random14chars:cachemiss] message
    """
    chars = string.ascii_letters + string.digits
    token = "".join(random.choices(chars, k=14))
    return f"[{token}:cachemiss] {message}"


# ============================================================================
# Question Set Loading
# ============================================================================

def loadQuestionSet(excelPath: str) -> List[Dict[str, Any]]:
    """
    Load questions from Excel file, sheet "Question Set".
    Columns: ID, Tier, Total Steps, Unique Tool Types, User Prompt, Follow-up 1, Follow-up 2, Expected Call Sequence, Computation / Notes
    Returns list of dicts with: id, tier, totalSteps, uniqueToolTypes, userPrompt, followUp1, followUp2, expectedCallSequence, computationNotes
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    if "Question Set" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Question Set' not found in {excelPath}")
    
    ws = wb["Question Set"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return []
    
    # Assume first row is header
    header = rows[0]
    questions = []
    
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        
        # Map columns by name (case-insensitive, flexible)
        rowDict = {}
        for i, val in enumerate(row):
            if i < len(header) and header[i]:
                colName = str(header[i]).strip().lower()
                rowDict[colName] = val
        
        # Extract fields
        qId = rowDict.get("id", "")
        tier = rowDict.get("tier", "")
        totalSteps = rowDict.get("total steps", "")
        uniqueToolTypes = rowDict.get("unique tool types", "")
        userPrompt = rowDict.get("user prompt", "")
        followUp1 = rowDict.get("follow-up 1", "")
        followUp2 = rowDict.get("follow-up 2", "")
        expectedCallSequence = rowDict.get("expected call sequence", "")
        computationNotes = rowDict.get("computation / notes", "")
        
        # Build followups list (non-empty only)
        followups = []
        if followUp1 and str(followUp1).strip():
            followups.append(str(followUp1).strip())
        if followUp2 and str(followUp2).strip():
            followups.append(str(followUp2).strip())
        
        questions.append({
            "id": str(qId).strip() if qId else "",
            "tier": str(tier).strip() if tier else "",
            "totalSteps": str(totalSteps).strip() if totalSteps else "",
            "uniqueToolTypes": str(uniqueToolTypes).strip() if uniqueToolTypes else "",
            "userPrompt": str(userPrompt).strip() if userPrompt else "",
            "followUp1": str(followUp1).strip() if followUp1 else "",
            "followUp2": str(followUp2).strip() if followUp2 else "",
            "expectedCallSequence": str(expectedCallSequence).strip() if expectedCallSequence else "",
            "computationNotes": str(computationNotes).strip() if computationNotes else "",
            "followups": followups,
        })
    
    return questions


# ============================================================================
# Flight-Assistant Client (Option A)
# ============================================================================

def callFlightAssistant(
    baseUrl: str,
    message: str,
    previousResponseId: Optional[str] = None,
    model: str = DEFAULT_MODEL,
    reasoning: str = "none"
) -> Dict[str, Any]:
    """
    POST to {baseUrl}/api/flight-assistant with { message, previousResponseId, model, reasoning }.
    Adds cache-busting token to message to prevent cached responses.
    Returns: { text, responseId, toolCalls, statusCode, error, usage, sentMessage }
    Each toolCall has: toolName, args, result
    usage has: inputTokens, outputTokens (if available from response)
    sentMessage: the actual message sent (with cache bust)
    """
    # Add cache buster to prevent API response caching
    messageWithCacheBust = addCacheBuster(message)
    
    url = f"{baseUrl.rstrip('/')}/api/flight-assistant"
    payload = {
        "message": messageWithCacheBust,
        "model": model,
        "reasoning": reasoning,
    }
    if previousResponseId:
        payload["previousResponseId"] = previousResponseId
    
    headers = {"Content-Type": "application/json"}
    
    try:
        if HAS_REQUESTS:
            resp = requests.post(url, json=payload, headers=headers, timeout=60)
            statusCode = resp.status_code
            if statusCode == 200:
                data = resp.json()
                # Extract usage if available
                usage = data.get("usage", {})
                return {
                    "text": data.get("text", ""),
                    "responseId": data.get("responseId"),
                    "toolCalls": data.get("toolCalls", []),
                    "statusCode": statusCode,
                    "error": None,
                    "usage": usage,
                    "sentMessage": messageWithCacheBust,
                }
            else:
                return {
                    "text": "",
                    "responseId": None,
                    "toolCalls": [],
                    "statusCode": statusCode,
                    "error": resp.text,
                    "usage": {},
                    "sentMessage": messageWithCacheBust,
                }
        else:
            # Use urllib
            req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                statusCode = resp.getcode()
                data = json.loads(resp.read().decode("utf-8"))
                usage = data.get("usage", {})
                return {
                    "text": data.get("text", ""),
                    "responseId": data.get("responseId"),
                    "toolCalls": data.get("toolCalls", []),
                    "statusCode": statusCode,
                    "error": None,
                    "usage": usage,
                    "sentMessage": messageWithCacheBust,
                }
    except Exception as e:
        return {
            "text": "",
            "responseId": None,
            "toolCalls": [],
            "statusCode": None,
            "error": str(e),
            "usage": {},
            "sentMessage": messageWithCacheBust,
        }


def runOneQuestion(
    question: Dict[str, Any],
    baseUrl: str,
    model: str = DEFAULT_MODEL,
    reasoning: str = "none"
) -> Dict[str, Any]:
    """
    Send initial user message (userPrompt), then any follow-ups.
    Returns: { turns: [...], totalUserMessages, totalToolCalls, perToolCounts, lastResponseId, totalInputTokens, totalOutputTokens }
    Each turn: { turnIndex, userMessage, sentMessage, responseText, toolCalls, inputTokens, outputTokens }
    """
    turns = []
    totalUserMessages = 0
    totalToolCalls = 0
    totalInputTokens = 0
    totalOutputTokens = 0
    perToolCounts = {
        "plotManagement": 0,
        "missionManagement": 0,
        "cameraSensors": 0,
        "executeFlightScript": 0,
    }
    
    previousResponseId = None
    
    # Initial prompt
    userMsg = question["userPrompt"]
    if not userMsg:
        return {
            "turns": turns,
            "totalUserMessages": 0,
            "totalToolCalls": 0,
            "perToolCounts": perToolCounts,
            "lastResponseId": None,
            "totalInputTokens": 0,
            "totalOutputTokens": 0,
        }
    
    totalUserMessages += 1
    result = callFlightAssistant(baseUrl, userMsg, previousResponseId, model, reasoning)
    previousResponseId = result.get("responseId")
    toolCalls = result.get("toolCalls", [])
    totalToolCalls += len(toolCalls)
    
    # Extract usage - try multiple field names for compatibility
    usage = result.get("usage", {})
    inputTokens = (
        usage.get("inputTokens") or 
        usage.get("input_tokens") or 
        usage.get("promptTokens") or 
        usage.get("prompt_tokens") or 
        0
    )
    outputTokens = (
        usage.get("outputTokens") or 
        usage.get("output_tokens") or 
        usage.get("completionTokens") or 
        usage.get("completion_tokens") or 
        0
    )
    totalInputTokens += inputTokens
    totalOutputTokens += outputTokens
    
    for tc in toolCalls:
        toolName = tc.get("toolName", "")
        if toolName in perToolCounts:
            perToolCounts[toolName] += 1
    
    turns.append({
        "turnIndex": 0,
        "userMessage": userMsg,
        "sentMessage": result.get("sentMessage", userMsg),
        "responseText": result.get("text", ""),
        "toolCalls": toolCalls,
        "inputTokens": inputTokens,
        "outputTokens": outputTokens,
    })
    
    # Follow-ups
    for i, followUpMsg in enumerate(question["followups"]):
        if not followUpMsg:
            continue
        totalUserMessages += 1
        result = callFlightAssistant(baseUrl, followUpMsg, previousResponseId, model, reasoning)
        previousResponseId = result.get("responseId")
        toolCalls = result.get("toolCalls", [])
        totalToolCalls += len(toolCalls)
        
        # Extract usage - try multiple field names for compatibility
        usage = result.get("usage", {})
        inputTokens = (
            usage.get("inputTokens") or 
            usage.get("input_tokens") or 
            usage.get("promptTokens") or 
            usage.get("prompt_tokens") or 
            0
        )
        outputTokens = (
            usage.get("outputTokens") or 
            usage.get("output_tokens") or 
            usage.get("completionTokens") or 
            usage.get("completion_tokens") or 
            0
        )
        totalInputTokens += inputTokens
        totalOutputTokens += outputTokens
        
        for tc in toolCalls:
            toolName = tc.get("toolName", "")
            if toolName in perToolCounts:
                perToolCounts[toolName] += 1
        
        turns.append({
            "turnIndex": i + 1,
            "userMessage": followUpMsg,
            "sentMessage": result.get("sentMessage", followUpMsg),
            "responseText": result.get("text", ""),
            "toolCalls": toolCalls,
            "inputTokens": inputTokens,
            "outputTokens": outputTokens,
        })
    
    return {
        "turns": turns,
        "totalUserMessages": totalUserMessages,
        "totalToolCalls": totalToolCalls,
        "perToolCounts": perToolCounts,
        "lastResponseId": previousResponseId,
        "totalInputTokens": totalInputTokens,
        "totalOutputTokens": totalOutputTokens,
    }


# ============================================================================
# Dummy Test Script
# ============================================================================

def callTestScript(runLog: Dict[str, Any]) -> None:
    """
    Dummy test script: does nothing.
    This is a no-op placeholder for a future real implementation.
    """
    pass


# ============================================================================
# Excel Logging
# ============================================================================

LOG_HEADERS = (
    "Run ID",
    "Question ID",
    "Model",
    "Reasoning",
    "Timestamp",
    "Total User Messages",
    "Total Tool Calls",
    "Input Tokens",
    "Output Tokens",
    "Time (s)",
    "plotManagement",
    "missionManagement",
    "cameraSensors",
    "executeFlightScript",
    "Interaction Sequence",
    "Turn Details (JSON)",
    "Test Script Result",
)

TURN_LOG_HEADERS = (
    "Run ID",
    "Question ID",
    "Model",
    "Reasoning",
    "Turn Index",
    "Message Type",
    "Original User Message",
    "Sent Message (with cache bust)",
    "API Response",
    "Input Tokens",
    "Output Tokens",
    "Tool Calls (Names)",
    "Tool Call Details (JSON)",
)

def formatInteractionSequence(turns: List[Dict[str, Any]]) -> str:
    """
    Format the sequence of user messages and tool calls in order.
    Example: "USER → toolA, toolB → USER → toolC"
    """
    sequence = []
    for i, turn in enumerate(turns):
        if i > 0:
            sequence.append(" → ")
        sequence.append("USER")
        toolCalls = turn.get("toolCalls", [])
        if toolCalls:
            toolNames = [tc.get("toolName", "unknown") for tc in toolCalls]
            sequence.append(" → ")
            sequence.append(", ".join(toolNames))
    return "".join(sequence)

def appendDetailedTurns(wb, runId: str, questionId: str, model: str, reasoning: str, turns: List[Dict[str, Any]]) -> None:
    """
    Append turn-by-turn details to a separate "Turn Details" sheet.
    Each row is one turn with user message, API response, tokens, and tool calls.
    """
    # Get or create Turn Details sheet
    if "Turn Details" in wb.sheetnames:
        ws = wb["Turn Details"]
    else:
        ws = wb.create_sheet("Turn Details")
        ws.append(TURN_LOG_HEADERS)
    
    for turn in turns:
        toolCalls = turn.get("toolCalls", [])
        toolNames = ", ".join([tc.get("toolName", "") for tc in toolCalls])
        toolCallsJson = json.dumps(
            [{"toolName": tc.get("toolName", ""), "args": tc.get("args", {})} for tc in toolCalls],
            ensure_ascii=False
        )
        
        row = (
            runId,
            questionId,
            model,
            reasoning,
            turn["turnIndex"],
            "USER",
            turn["userMessage"],
            turn.get("sentMessage", turn["userMessage"]),
            turn["responseText"],
            turn.get("inputTokens", 0),
            turn.get("outputTokens", 0),
            toolNames,
            toolCallsJson,
        )
        ws.append(row)

def appendRunToExcel(
    logPath: str,
    runId: str,
    question: Dict[str, Any],
    runLog: Dict[str, Any],
    model: str,
    reasoning: str,
    elapsed: float
) -> None:
    """
    Append one row per run to Excel log.
    Includes detailed turn-by-turn information with questions, responses, tokens, and ordered tool/user calls.
    Also adds detailed turn logs to a separate sheet.
    """
    if not HAS_OPENPYXL:
        print("openpyxl not available; skipping Excel log.", file=sys.stderr)
        return
    
    if os.path.isfile(logPath):
        wb = load_workbook(logPath)
        if wb.active.title != "Summary":
            # Rename first sheet to Summary if not already
            wb.active.title = "Summary"
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(LOG_HEADERS)
    
    perToolCounts = runLog["perToolCounts"]
    turns = runLog.get("turns", [])
    
    # Format interaction sequence
    interactionSeq = formatInteractionSequence(turns)
    
    # Format turn details as JSON for detailed inspection
    turnDetails = []
    for turn in turns:
        turnDetail = {
            "turn": turn["turnIndex"],
            "userMessage": turn["userMessage"],
            "responseText": turn["responseText"][:200] + "..." if len(turn["responseText"]) > 200 else turn["responseText"],
            "inputTokens": turn.get("inputTokens", 0),
            "outputTokens": turn.get("outputTokens", 0),
            "toolCalls": [
                {
                    "toolName": tc.get("toolName", ""),
                    "args": tc.get("args", {}),
                }
                for tc in turn.get("toolCalls", [])
            ]
        }
        turnDetails.append(turnDetail)
    
    turnDetailsJson = json.dumps(turnDetails, indent=2, ensure_ascii=False)
    
    row = (
        runId,
        question["id"],
        model,
        reasoning,
        datetime.now().isoformat(),
        runLog["totalUserMessages"],
        runLog["totalToolCalls"],
        runLog.get("totalInputTokens", 0),
        runLog.get("totalOutputTokens", 0),
        f"{elapsed:.2f}",
        perToolCounts["plotManagement"],
        perToolCounts["missionManagement"],
        perToolCounts["cameraSensors"],
        perToolCounts["executeFlightScript"],
        interactionSeq,
        turnDetailsJson,
        "",  # Test Script Result (empty for now)
    )
    ws.append(row)
    
    # Add detailed turn logs to separate sheet
    appendDetailedTurns(wb, runId, question["id"], model, reasoning, turns)
    
    wb.save(logPath)


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Pass^k Tool-Call Experiment Runner (Option A)")
    parser.add_argument("--base-url", default="http://localhost:3000", help="Base URL for flight-assistant API")
    parser.add_argument("--question-set-path", default=None, help="Path to question set Excel file")
    parser.add_argument("--log-path", default=None, help="Path to output Excel log file")
    parser.add_argument("--model", default=None, help="Model to use (default: run all configurations)")
    parser.add_argument("--reasoning", default=None, help="Reasoning effort: none, high (default: run all configurations)")
    parser.add_argument("--runs-per-config", type=int, default=1, help="Number of runs per configuration (default: 1)")
    args = parser.parse_args()
    
    # Default paths
    scriptDir = os.path.dirname(os.path.abspath(__file__))
    questionSetPath = args.question_set_path or os.path.join(scriptDir, "data", "pass_k_question_set.xlsx")
    logPath = args.log_path or os.path.join(scriptDir, "pass_k_tool_call_log.xlsx")
    
    if not os.path.isfile(questionSetPath):
        print(f"Question set not found: {questionSetPath}", file=sys.stderr)
        sys.exit(1)
    
    # Load questions
    print(f"Loading questions from: {questionSetPath}")
    questions = loadQuestionSet(questionSetPath)
    print(f"Loaded {len(questions)} questions.")
    
    if not questions:
        print("No questions to run.", file=sys.stderr)
        sys.exit(0)
    
    # Define model configurations
    if args.model and args.reasoning:
        # Single configuration specified
        configurations = [(args.model, args.reasoning)]
    else:
        # Run all configurations: gpt-5.2 (none, high), gpt-5.4 (none, high)
        configurations = [
            ("gpt-5.2", "none"),
            ("gpt-5.2", "high"),
            ("gpt-5.4", "none"),
            ("gpt-5.4", "high"),
        ]
    
    # Run experiments
    print(f"Base URL: {args.base_url}")
    print(f"Log path: {logPath}")
    print(f"Configurations: {len(configurations)}, runs per config: {args.runs_per_config}")
    print()
    
    totalRuns = 0
    
    for configIdx, (model, reasoning) in enumerate(configurations):
        configName = f"{model}-reasoning-{reasoning}"
        print(f"=== Configuration {configIdx + 1}/{len(configurations)}: {configName} ===")
        
        for runNum in range(args.runs_per_config):
            for qIdx, question in enumerate(questions):
                totalRuns += 1
                runId = f"{configName}-run{runNum + 1}-q{qIdx + 1}"
                qId = question["id"] or f"Q{qIdx + 1}"
                print(f"[{runId}] Running question {qId}...")
                
                startTime = time.perf_counter()
                runLog = runOneQuestion(question, args.base_url, model, reasoning)
                elapsed = time.perf_counter() - startTime
                
                print(f"  Model: {model}, Reasoning: {reasoning}")
                print(f"  User messages: {runLog['totalUserMessages']}, Tool calls: {runLog['totalToolCalls']}")
                print(f"  Tokens: in={runLog.get('totalInputTokens', 0)}, out={runLog.get('totalOutputTokens', 0)}")
                print(f"  Per-tool: {runLog['perToolCounts']}")
                print(f"  Interaction sequence: {formatInteractionSequence(runLog['turns'])}")
                
                # Show turn details
                for turn in runLog['turns']:
                    print(f"    Turn {turn['turnIndex']}: USER")
                    userMsg = turn['userMessage']
                    print(f"      → Question: {userMsg[:80]}..." if len(userMsg) > 80 else f"      → Question: {userMsg}")
                    respText = turn['responseText']
                    print(f"      → Response: {respText[:80]}..." if len(respText) > 80 else f"      → Response: {respText}")
                    print(f"      → Tokens: in={turn.get('inputTokens', 0)}, out={turn.get('outputTokens', 0)}")
                    toolCalls = turn.get('toolCalls', [])
                    if toolCalls:
                        print(f"      → Tool calls: {', '.join([tc.get('toolName', 'unknown') for tc in toolCalls])}")
                
                print(f"  Time: {elapsed:.2f}s")
                
                # Call dummy test script
                callTestScript(runLog)
                
                # Log to Excel
                appendRunToExcel(logPath, runId, question, runLog, model, reasoning, elapsed)
                print()
        
        print()
    
    print(f"All runs completed. Total runs: {totalRuns}")
    print(f"Log saved to: {logPath}")


if __name__ == "__main__":
    main()
