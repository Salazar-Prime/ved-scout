#!/usr/bin/env python3
"""
Plot pass^k (pass-hat-k) curves by task difficulty and model configuration.

pass^k = mean(p_i ^ k) across questions i in a difficulty group,
where p_i = (# passing runs) / (# total runs) for question i.

This is the probability that ALL k independent runs pass — a stricter
metric than pass@k which only requires at least one pass.
"""

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "output" / "Exp-Pass-k-calc-evaluated-Final.xlsx"

DIFFICULTY_MAP = {"E": "Easy", "M": "Medium", "H": "Hard"}
DIFFICULTY_ORDER = ["E", "M", "H"]
MAX_K = 5


def loadData(xlsxPath):
    wb = openpyxl.load_workbook(xlsxPath, data_only=True)
    ws = wb["Summary"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def computePerQuestionPassRates(rows):
    """
    Returns:
        modelConfigs: sorted list of model config labels
        passRates: dict (modelConfig, questionId) -> p (estimated pass probability)
    """
    runResults = defaultdict(list)
    for r in rows:
        model = r["Model"]
        reasoning = r["Reasoning"]
        qid = r["Question ID"]
        grade = str(r["Human Grade"]).strip().lower()
        configLabel = f"{model}\n(reasoning={reasoning})"
        runResults[(configLabel, qid)].append(1 if grade == "pass" else 0)

    passRates = {}
    for (config, qid), outcomes in runResults.items():
        passRates[(config, qid)] = sum(outcomes) / len(outcomes)

    modelConfigs = sorted(set(cfg for cfg, _ in passRates.keys()))
    return modelConfigs, passRates


def passHatK(ps, k):
    """Probability ALL k runs pass: mean(p_i^k)."""
    return np.mean(np.array(ps) ** k)


def passAtK(ps, k):
    """Probability at least 1 of k runs passes: mean(1 - (1-p_i)^k)."""
    return np.mean(1.0 - (1.0 - np.array(ps)) ** k)


def buildCurves(modelConfigs, passRates):
    """
    Build pass^k and pass@k curves for each (modelConfig, difficulty)
    and also for all tasks combined (key "All").

    Returns:
        ks: array [1..MAX_K]
        hatCurves: dict (config, difficulty|"All") -> array of pass^k values
        atCurves:  dict (config, difficulty|"All") -> array of pass@k values
    """
    ks = np.arange(1, MAX_K + 1)
    hatCurves = {}
    atCurves = {}

    for config in modelConfigs:
        for d in DIFFICULTY_ORDER:
            ps = [passRates[(config, qid)]
                  for (c, qid) in passRates if c == config and qid.startswith(d)]
            hatCurves[(config, d)] = np.array([passHatK(ps, k) for k in ks])
            atCurves[(config, d)] = np.array([passAtK(ps, k) for k in ks])

        allPs = [passRates[(c, qid)] for (c, qid) in passRates if c == config]
        hatCurves[(config, "All")] = np.array([passHatK(allPs, k) for k in ks])
        atCurves[(config, "All")] = np.array([passAtK(allPs, k) for k in ks])

    return ks, hatCurves, atCurves


def plotKSweepCombined(modelConfigs, ks, hatCurves, atCurves):
    """Single plot with all tasks combined: pass^k (solid) and pass@k (dashed) vs k."""
    colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]
    fig, ax = plt.subplots(figsize=(9, 6))

    for i, config in enumerate(modelConfigs):
        c = colors[i % len(colors)]
        configOneLine = config.replace("\n", " ")
        ax.plot(ks, hatCurves[(config, "All")], "o-", color=c,
                label=f"pass^k  {configOneLine}", linewidth=2.2, markersize=6)
        ax.plot(ks, atCurves[(config, "All")], "s--", color=c,
                label=f"pass@k {configOneLine}", linewidth=1.5, markersize=5, alpha=0.55)

    ax.set_title("pass^k vs pass@k — All Tasks Combined (n=30)",
                 fontsize=14, fontweight="bold")
    ax.set_xlabel("k (repeated runs)", fontsize=12, fontweight="bold")
    ax.set_ylabel("Probability", fontsize=12, fontweight="bold")
    ax.set_xticks(ks)
    ax.set_ylim(-0.02, 1.05)
    ax.legend(fontsize=9, framealpha=0.9, loc="center left", bbox_to_anchor=(0.01, 0.5))
    ax.grid(True, linestyle=":", linewidth=0.7, alpha=0.6)
    ax.set_axisbelow(True)

    fig.tight_layout()
    return fig


def plotKSweepByDifficulty(modelConfigs, ks, hatCurves, atCurves):
    """One subplot per difficulty: pass^k (solid) and pass@k (dashed) vs k."""
    colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]
    fig, axes = plt.subplots(1, 3, figsize=(16, 5), sharey=True)

    for ax, d in zip(axes, DIFFICULTY_ORDER):
        for i, config in enumerate(modelConfigs):
            c = colors[i % len(colors)]
            ax.plot(ks, hatCurves[(config, d)], "o-", color=c, label=f"pass^k  {config}", linewidth=2, markersize=5)
            ax.plot(ks, atCurves[(config, d)], "s--", color=c, label=f"pass@k {config}", linewidth=1.5, markersize=4, alpha=0.6)

        ax.set_title(f"{DIFFICULTY_MAP[d]} Tasks", fontsize=13, fontweight="bold")
        ax.set_xlabel("k (repeated runs)", fontsize=11)
        ax.set_xticks(ks)
        ax.set_ylim(-0.02, 1.05)
        ax.grid(True, linestyle=":", linewidth=0.7, alpha=0.6)
        ax.set_axisbelow(True)

    axes[0].set_ylabel("Probability", fontsize=12, fontweight="bold")

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncol=3, fontsize=8,
               bbox_to_anchor=(0.5, -0.08), framealpha=0.9)

    fig.suptitle("pass^k vs pass@k — by Difficulty & Model Configuration",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    return fig


def plotPassHatKBar(modelConfigs, hatCurves, k=1):
    """Bar chart: pass^k grouped by difficulty with one bar per model config."""
    kIdx = k - 1
    nGroups = len(DIFFICULTY_ORDER)
    nBars = len(modelConfigs)
    barWidth = 0.22
    x = np.arange(nGroups)
    colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]

    fig, ax = plt.subplots(figsize=(10, 6))

    for i, config in enumerate(modelConfigs):
        heights = [hatCurves[(config, d)][kIdx] for d in DIFFICULTY_ORDER]
        offsets = x + (i - (nBars - 1) / 2) * barWidth
        bars = ax.bar(offsets, heights, barWidth, label=config,
                      color=colors[i % len(colors)], edgecolor="white", linewidth=0.5)
        for bar, h in zip(bars, heights):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.01,
                    f"{h:.2f}", ha="center", va="bottom", fontsize=9, fontweight="bold")

    ax.set_xlabel("Task Difficulty", fontsize=12, fontweight="bold")
    ax.set_ylabel(f"pass^{k}", fontsize=12, fontweight="bold")
    ax.set_title(f"pass^{k} by Difficulty & Model Configuration", fontsize=14, fontweight="bold")
    ax.set_xticks(x)
    ax.set_xticklabels([DIFFICULTY_MAP[d] for d in DIFFICULTY_ORDER], fontsize=11)
    ax.set_ylim(0, 1.12)
    ax.legend(loc="upper right", fontsize=9, framealpha=0.9)
    ax.grid(axis="y", alpha=0.3, linestyle="--")
    ax.set_axisbelow(True)

    fig.tight_layout()
    return fig


def plotRuntimeBoxplot(rows):
    """Box plot of per-task runtime (s) across all runs, one box per model config."""
    colors = ["#4C72B0", "#DD8452", "#55A868", "#C44E52"]

    configToTimes = defaultdict(list)
    for r in rows:
        configLabel = f"{r['Model']}\n(reasoning={r['Reasoning']})"
        if r["Time (s)"] is not None:
            configToTimes[configLabel].append(float(r["Time (s)"]))

    configs = sorted(configToTimes.keys())
    data = [configToTimes[c] for c in configs]

    fig, ax = plt.subplots(figsize=(9, 6))

    bp = ax.boxplot(data, patch_artist=True, widths=0.5,
                    medianprops=dict(color="black", linewidth=1.5),
                    whiskerprops=dict(linewidth=1.2),
                    capprops=dict(linewidth=1.2),
                    flierprops=dict(marker="o", markersize=4, alpha=0.5))

    for patch, c in zip(bp["boxes"], colors):
        patch.set_facecolor(c)
        patch.set_alpha(0.7)

    for i, (c, times) in enumerate(zip(configs, data)):
        median = np.median(times)
        mean = np.mean(times)
        ax.text(i + 1, median + 0.3, f"med={median:.1f}s",
                ha="center", va="bottom", fontsize=8, fontweight="bold")
        ax.plot(i + 1, mean, "D", color="white", markeredgecolor="black",
                markersize=6, zorder=5)

    ax.set_xticklabels([c.replace("\n", " ") for c in configs], fontsize=10)
    ax.set_ylabel("Runtime per Task (s)", fontsize=12, fontweight="bold")
    ax.set_title("Runtime Distribution by Model Configuration", fontsize=14, fontweight="bold")
    ax.grid(axis="y", alpha=0.3, linestyle="--")
    ax.set_axisbelow(True)

    from matplotlib.lines import Line2D
    legend = [Line2D([0], [0], marker="D", color="white", markeredgecolor="black",
                     markersize=6, label="Mean", linestyle="None")]
    ax.legend(handles=legend, loc="upper right", fontsize=9)

    fig.tight_layout()
    return fig


def printRunStats(rows):
    """Print average time, tokens, and cost — overall and broken down by run number."""
    configToRows = defaultdict(list)
    for r in rows:
        configLabel = f"{r['Model']} (reasoning={r['Reasoning']})"
        configToRows[configLabel].append(r)

    def safeFloats(rList, key):
        return [float(r[key]) for r in rList if r[key] is not None]

    for config in sorted(configToRows.keys()):
        cRows = configToRows[config]
        print(f"\n{config}:")

        runBuckets = defaultdict(list)
        for r in cRows:
            runBuckets[r["Run #"]].append(r)

        print(f"  {'Run':<8} {'Avg Time(s)':>12} {'Avg InTok':>10} {'Avg OutTok':>11} {'Avg Cost($)':>12} {'Count':>6}")
        print(f"  {'-'*8} {'-'*12} {'-'*10} {'-'*11} {'-'*12} {'-'*6}")

        for run in sorted(runBuckets.keys()):
            rRows = runBuckets[run]
            avgTime = np.mean(safeFloats(rRows, "Time (s)"))
            avgIn = np.mean(safeFloats(rRows, "Input Tokens"))
            avgOut = np.mean(safeFloats(rRows, "Output Tokens"))
            avgCost = np.mean(safeFloats(rRows, "Est. Cost ($)"))
            print(f"  {run:<8} {avgTime:>12.2f} {avgIn:>10.0f} {avgOut:>11.0f} {avgCost:>12.4f} {len(rRows):>6}")

        avgTime = np.mean(safeFloats(cRows, "Time (s)"))
        avgIn = np.mean(safeFloats(cRows, "Input Tokens"))
        avgOut = np.mean(safeFloats(cRows, "Output Tokens"))
        avgCost = np.mean(safeFloats(cRows, "Est. Cost ($)"))
        totalCost = sum(safeFloats(cRows, "Est. Cost ($)"))
        totalTime = sum(safeFloats(cRows, "Time (s)"))
        print(f"  {'-'*8} {'-'*12} {'-'*10} {'-'*11} {'-'*12} {'-'*6}")
        print(f"  {'Overall':<8} {avgTime:>12.2f} {avgIn:>10.0f} {avgOut:>11.0f} {avgCost:>12.4f} {len(cRows):>6}")
        print(f"  {'':8} {'Total Time:':>12} {totalTime:>10.1f}s {'':11} {'Total:':>12} ${totalCost:.4f}")

    allCosts = safeFloats(rows, "Est. Cost ($)")
    allTimes = safeFloats(rows, "Time (s)")
    allIn = safeFloats(rows, "Input Tokens")
    allOut = safeFloats(rows, "Output Tokens")
    grandTotal = sum(allCosts)
    grandTime = sum(allTimes)
    avgIn = np.mean(allIn)
    avgOut = np.mean(allOut)
    ratio = avgIn / avgOut if avgOut > 0 else float("inf")

    print(f"\n{'='*70}")
    print(f"  Grand Total Cost (all configs, all runs): ${grandTotal:.4f}")
    print(f"  Grand Total Time (all configs, all runs): {grandTime:.1f}s ({grandTime/60:.1f}m)")
    print(f"  Total Runs: {len(rows)}")
    print(f"  Avg Input Tokens:  {avgIn:,.0f}")
    print(f"  Avg Output Tokens: {avgOut:,.0f}")
    print(f"  Input/Output Ratio: {ratio:.2f}:1")
    print(f"{'='*70}")

    print(f"\n  {'Config':<35} {'Avg In':>8} {'Avg Out':>9} {'Ratio':>8}")
    print(f"  {'-'*35} {'-'*8} {'-'*9} {'-'*8}")
    for config in sorted(configToRows.keys()):
        cRows = configToRows[config]
        cIn = np.mean(safeFloats(cRows, "Input Tokens"))
        cOut = np.mean(safeFloats(cRows, "Output Tokens"))
        cRatio = cIn / cOut if cOut > 0 else float("inf")
        print(f"  {config:<35} {cIn:>8,.0f} {cOut:>9,.0f} {cRatio:>7.2f}:1")


def main():
    rows = loadData(XLSX_PATH)
    modelConfigs, passRates = computePerQuestionPassRates(rows)
    ks, hatCurves, atCurves = buildCurves(modelConfigs, passRates)

    outDir = Path(__file__).parent / "output"

    # Print tables
    print("=== pass^k values ===")
    for config in modelConfigs:
        configOneLine = config.replace("\n", " ")
        print(f"\n{configOneLine}:")
        for d in DIFFICULTY_ORDER:
            vals = "  ".join(f"k={k}: {v:.4f}" for k, v in zip(ks, hatCurves[(config, d)]))
            print(f"  {DIFFICULTY_MAP[d]:>8}: {vals}")
        vals = "  ".join(f"k={k}: {v:.4f}" for k, v in zip(ks, hatCurves[(config, "All")]))
        print(f"  {'All':>8}: {vals}")

    print("\n=== Per-Question pass rates (p_i) ===")
    for config in modelConfigs:
        configOneLine = config.replace("\n", " ")
        print(f"\n{configOneLine}:")
        for d in DIFFICULTY_ORDER:
            qids = sorted([qid for (c, qid) in passRates if c == config and qid.startswith(d)])
            for qid in qids:
                p = passRates[(config, qid)]
                print(f"  {qid:>4}: p={p:.2f}")

    print("\n=== Run Statistics (Time, Tokens, Cost) ===")
    printRunStats(rows)

    # Plot 1a: k-sweep — all tasks combined
    fig1a = plotKSweepCombined(modelConfigs, ks, hatCurves, atCurves)
    path1a = outDir / "pass_hat_k_sweep.png"
    fig1a.savefig(path1a, dpi=150, bbox_inches="tight")
    print(f"\nk-sweep (combined) saved to: {path1a}")

    # Plot 1b: k-sweep — per difficulty
    fig1b = plotKSweepByDifficulty(modelConfigs, ks, hatCurves, atCurves)
    path1b = outDir / "pass_hat_k_sweep_by_difficulty.png"
    fig1b.savefig(path1b, dpi=150, bbox_inches="tight")
    print(f"k-sweep (by difficulty) saved to: {path1b}")

    # Plot 2: pass^k bar charts for k=1 and k=5
    for k in [1, 5]:
        fig = plotPassHatKBar(modelConfigs, hatCurves, k=k)
        path = outDir / f"pass_hat_{k}_bar.png"
        fig.savefig(path, dpi=150, bbox_inches="tight")
        print(f"pass^{k} bar chart saved to: {path}")

    # Plot 3: runtime box plot
    fig3 = plotRuntimeBoxplot(rows)
    path3 = outDir / "runtime_boxplot.png"
    fig3.savefig(path3, dpi=150, bbox_inches="tight")
    print(f"Runtime box plot saved to: {path3}")

    plt.show()


if __name__ == "__main__":
    main()
