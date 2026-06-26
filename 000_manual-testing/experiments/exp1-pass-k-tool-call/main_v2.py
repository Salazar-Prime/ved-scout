#!/usr/bin/env python3
"""
Pass^k Tool-Call Experiment Runner (Refactored Architecture)
Python handles OpenAI API calls directly for full control.
Next.js handles Firebase tool execution only.
"""
import argparse
import json
import os
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

from openaiClient import createOpenAiClient, OpenAIClient
from toolsClient import createToolsClient, ToolsClient
from config import (
    DEFAULT_MODEL,
    DEFAULT_REASONING,
    SYSTEM_PROMPT,
    TOOL_SCHEMAS,
    MODEL_PRICING,
)

# Excel logging
try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)


# ============================================================================
# Cost Estimation
# ============================================================================

def estimateCost(inputTokens: int, outputTokens: int, model: str) -> Optional[float]:
    """
    Estimate cost in USD from token counts.
    Returns float or None if model unknown.
    """
    if (inputTokens or 0) == 0 and (outputTokens or 0) == 0:
        return 0.0
    
    # Extract base model name (handle cases like "gpt-5.2-2024-01-01")
    baseModel = model.split("/")[-1].split("-202")[0] if model else ""
    prices = MODEL_PRICING.get(baseModel) or MODEL_PRICING.get(model)
    
    if not prices:
        return None
    
    inputPerM, outputPerM = prices
    return (inputTokens or 0) * inputPerM / 1e6 + (outputTokens or 0) * outputPerM / 1e6


# ============================================================================
# Question Set Loading
# ============================================================================

def loadQuestionSet(excelPath: str) -> List[Dict[str, Any]]:
    """
    Load questions from Excel file, sheet "Question Set".
    Returns list of dicts with: id, tier, totalSteps, uniqueToolTypes, userPrompt, followUp1, followUp2, expectedCallSequence, computationNotes
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    if "Question Set" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Question Set' not found in {excelPath}")
    
    ws = wb["Question Set"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return []
    
    header = rows[0]
    questions = []
    
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        
        rowDict = {}
        for i, val in enumerate(row):
            if i < len(header) and header[i]:
                colName = str(header[i]).strip().lower()
                rowDict[colName] = val
        
        qId = rowDict.get("id", "")
        tier = rowDict.get("tier", "")
        totalSteps = rowDict.get("total steps", "")
        uniqueToolTypes = rowDict.get("unique tool types", "")
        userPrompt = rowDict.get("user prompt", "")
        followUp1 = rowDict.get("follow-up 1", "")
        followUp2 = rowDict.get("follow-up 2", "")
        expectedCallSequence = rowDict.get("expected call sequence", "")
        computationNotes = rowDict.get("computation / notes", "")
        
        followups = []
        if followUp1 and str(followUp1).strip():
            followups.append(str(followUp1).strip())
        if followUp2 and str(followUp2).strip():
            followups.append(str(followUp2).strip())
        
        questions.append({
            "id": str(qId).strip() if qId else "",
            "tier": str(tier).strip() if tier else "",
            "totalSteps": str(totalSteps).strip() if totalSteps else "",
            "uniqueToolTypes": str(uniqueToolTypes).strip() if uniqueToolTypes else "",
            "userPrompt": str(userPrompt).strip() if userPrompt else "",
            "followUp1": str(followUp1).strip() if followUp1 else "",
            "followUp2": str(followUp2).strip() if followUp2 else "",
            "expectedCallSequence": str(expectedCallSequence).strip() if expectedCallSequence else "",
            "computationNotes": str(computationNotes).strip() if computationNotes else "",
            "followups": followups,
        })
    
    return questions


# ============================================================================
# Agent Runner (OpenAI + Tools)
# ============================================================================

class AgentRunner:
    """
    Runs agent with OpenAI for reasoning and Next.js for tool execution.
    Manages conversation state and tool execution loop.
    """
    
    def __init__(
        self,
        openaiClient: OpenAIClient,
        toolsClient: ToolsClient,
        systemPrompt: str = SYSTEM_PROMPT,
    ):
        self.openaiClient = openaiClient
        self.toolsClient = toolsClient
        self.systemPrompt = systemPrompt
        self.messages: List[Dict[str, Any]] = []
        self.conversationHistory: List[Dict[str, Any]] = []
    
    def runConversation(
        self,
        userMessage: str,
        maxIterations: int = 10,
    ) -> Dict[str, Any]:
        """
        Run one conversation turn with user message.
        Handles tool calling loop automatically.
        
        Returns:
            {
                "finalResponse": str,
                "toolCalls": [...],
                "usage": {...},
                "iterations": int,
            }
        """
        # Initialize messages if empty
        if not self.messages:
            self.messages = [
                {"role": "system", "content": self.systemPrompt}
            ]
        
        # Add user message
        self.messages.append({
            "role": "user",
            "content": userMessage,
        })
        
        allToolCalls = []
        totalUsage = {
            "promptTokens": 0,
            "completionTokens": 0,
            "totalTokens": 0,
        }
        
        iterations = 0
        finalResponse = ""
        
        # Tool calling loop
        for iteration in range(maxIterations):
            iterations += 1
            
            # Call OpenAI
            response = self.openaiClient.chatCompletion(
                messages=self.messages,
                tools=TOOL_SCHEMAS,
                toolChoice="auto",
                reasoning=DEFAULT_REASONING,
                cacheBust=True,
            )
            
            # Accumulate usage
            usage = response.get("usage", {})
            totalUsage["promptTokens"] += usage.get("promptTokens", 0)
            totalUsage["completionTokens"] += usage.get("completionTokens", 0)
            totalUsage["totalTokens"] += usage.get("totalTokens", 0)
            
            if response.get("error"):
                finalResponse = f"Error: {response['error']}"
                break
            
            content = response.get("content", "")
            toolCalls = response.get("toolCalls", [])
            
            # Add assistant message to conversation
            assistantMessage: Dict[str, Any] = {
                "role": "assistant",
                "content": content,
            }
            if toolCalls:
                assistantMessage["tool_calls"] = toolCalls
            
            self.messages.append(assistantMessage)
            
            # If no tool calls, we're done
            if not toolCalls:
                finalResponse = content
                break
            
            # Execute tools
            toolResults = self.toolsClient.executeToolCalls(toolCalls)
            
            # Add tool results to conversation
            for tr in toolResults:
                allToolCalls.append({
                    "toolName": tr["toolName"],
                    "arguments": tr["arguments"],
                    "result": tr["result"],
                })
                
                self.messages.append({
                    "role": "tool",
                    "tool_call_id": tr["toolCallId"],
                    "content": json.dumps(tr["result"]),
                })
        
        return {
            "finalResponse": finalResponse,
            "toolCalls": allToolCalls,
            "usage": totalUsage,
            "iterations": iterations,
        }


# ============================================================================
# Question Runner
# ============================================================================

def runOneQuestion(
    question: Dict[str, Any],
    openaiClient: OpenAIClient,
    toolsClient: ToolsClient,
) -> Dict[str, Any]:
    """
    Run one question with initial prompt and follow-ups.
    Returns: { turns: [...], totalUserMessages, totalToolCalls, perToolCounts, totalInputTokens, totalOutputTokens }
    """
    agent = AgentRunner(openaiClient, toolsClient)
    
    turns = []
    totalUserMessages = 0
    totalToolCalls = 0
    totalInputTokens = 0
    totalOutputTokens = 0
    perToolCounts = {
        "plotManagement": 0,
        "missionManagement": 0,
        "cameraSensors": 0,
        "executeFlightScript": 0,
    }
    
    # Initial prompt
    userMsg = question["userPrompt"]
    if not userMsg:
        return {
            "turns": turns,
            "totalUserMessages": 0,
            "totalToolCalls": 0,
            "perToolCounts": perToolCounts,
            "totalInputTokens": 0,
            "totalOutputTokens": 0,
        }
    
    totalUserMessages += 1
    result = agent.runConversation(userMsg, maxIterations=10)
    
    toolCalls = result["toolCalls"]
    totalToolCalls += len(toolCalls)
    
    usage = result["usage"]
    inputTokens = usage.get("promptTokens", 0)
    outputTokens = usage.get("completionTokens", 0)
    totalInputTokens += inputTokens
    totalOutputTokens += outputTokens
    
    for tc in toolCalls:
        toolName = tc.get("toolName", "")
        if toolName in perToolCounts:
            perToolCounts[toolName] += 1
    
    turns.append({
        "turnIndex": 0,
        "userMessage": userMsg,
        "sentMessage": userMsg,  # Already cache-busted internally
        "responseText": result["finalResponse"],
        "toolCalls": toolCalls,
        "inputTokens": inputTokens,
        "outputTokens": outputTokens,
    })
    
    # Follow-ups
    for i, followUpMsg in enumerate(question["followups"]):
        if not followUpMsg:
            continue
        
        totalUserMessages += 1
        result = agent.runConversation(followUpMsg, maxIterations=10)
        
        toolCalls = result["toolCalls"]
        totalToolCalls += len(toolCalls)
        
        usage = result["usage"]
        inputTokens = usage.get("promptTokens", 0)
        outputTokens = usage.get("completionTokens", 0)
        totalInputTokens += inputTokens
        totalOutputTokens += outputTokens
        
        for tc in toolCalls:
            toolName = tc.get("toolName", "")
            if toolName in perToolCounts:
                perToolCounts[toolName] += 1
        
        turns.append({
            "turnIndex": i + 1,
            "userMessage": followUpMsg,
            "sentMessage": followUpMsg,
            "responseText": result["finalResponse"],
            "toolCalls": toolCalls,
            "inputTokens": inputTokens,
            "outputTokens": outputTokens,
        })
    
    return {
        "turns": turns,
        "totalUserMessages": totalUserMessages,
        "totalToolCalls": totalToolCalls,
        "perToolCounts": perToolCounts,
        "totalInputTokens": totalInputTokens,
        "totalOutputTokens": totalOutputTokens,
    }


# ============================================================================
# Dummy Test Script
# ============================================================================

def callTestScript(runLog: Dict[str, Any]) -> None:
    """Dummy test script: does nothing."""
    pass


# ============================================================================
# Excel Logging
# ============================================================================

LOG_HEADERS = (
    "Run ID",
    "Question ID",
    "Model",
    "Reasoning",
    "Timestamp",
    "Total User Messages",
    "Total Tool Calls",
    "Input Tokens",
    "Output Tokens",
    "Est. Cost ($)",
    "Time (s)",
    "plotManagement",
    "missionManagement",
    "cameraSensors",
    "executeFlightScript",
    "Interaction Sequence",
    "Turn Details (JSON)",
    "Test Script Result",
)

TURN_LOG_HEADERS = (
    "Run ID",
    "Question ID",
    "Model",
    "Reasoning",
    "Turn Index",
    "Message Type",
    "Original User Message",
    "Sent Message (with cache bust)",
    "API Response",
    "Input Tokens",
    "Output Tokens",
    "Tool Calls (Names)",
    "Tool Call Details (JSON)",
)

def formatInteractionSequence(turns: List[Dict[str, Any]]) -> str:
    """Format the sequence of user messages and tool calls in order."""
    sequence = []
    for i, turn in enumerate(turns):
        if i > 0:
            sequence.append(" → ")
        sequence.append("USER")
        toolCalls = turn.get("toolCalls", [])
        if toolCalls:
            toolNames = [tc.get("toolName", "unknown") for tc in toolCalls]
            sequence.append(" → ")
            sequence.append(", ".join(toolNames))
    return "".join(sequence)

def appendDetailedTurns(wb, runId: str, questionId: str, model: str, reasoning: str, turns: List[Dict[str, Any]]) -> None:
    """Append turn-by-turn details to a separate "Turn Details" sheet."""
    if "Turn Details" in wb.sheetnames:
        ws = wb["Turn Details"]
    else:
        ws = wb.create_sheet("Turn Details")
        ws.append(TURN_LOG_HEADERS)
    
    for turn in turns:
        toolCalls = turn.get("toolCalls", [])
        toolNames = ", ".join([tc.get("toolName", "") for tc in toolCalls])
        toolCallsJson = json.dumps(
            [{"toolName": tc.get("toolName", ""), "args": tc.get("arguments", {})} for tc in toolCalls],
            ensure_ascii=False
        )
        
        row = (
            runId,
            questionId,
            model,
            reasoning,
            turn["turnIndex"],
            "USER",
            turn["userMessage"],
            turn.get("sentMessage", turn["userMessage"]),
            turn["responseText"],
            turn.get("inputTokens", 0),
            turn.get("outputTokens", 0),
            toolNames,
            toolCallsJson,
        )
        ws.append(row)

def appendRunToExcel(
    logPath: str,
    runId: str,
    question: Dict[str, Any],
    runLog: Dict[str, Any],
    model: str,
    reasoning: str,
    elapsed: float
) -> None:
    """Append one row per run to Excel log."""
    if not HAS_OPENPYXL:
        print("openpyxl not available; skipping Excel log.", file=sys.stderr)
        return
    
    if os.path.isfile(logPath):
        wb = load_workbook(logPath)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
        else:
            ws = wb.create_sheet("Summary", 0)
            ws.append(LOG_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(LOG_HEADERS)
    
    perToolCounts = runLog["perToolCounts"]
    turns = runLog.get("turns", [])
    
    interactionSeq = formatInteractionSequence(turns)
    
    # Build detailed turn information (no clipping - full text preserved)
    turnDetails = []
    for turn in turns:
        turnDetail = {
            "turn": turn["turnIndex"],
            "userMessage": turn["userMessage"],
            "responseText": turn["responseText"],  # Full text, no clipping
            "inputTokens": turn.get("inputTokens", 0),
            "outputTokens": turn.get("outputTokens", 0),
            "toolCalls": [
                {
                    "toolName": tc.get("toolName", ""),
                    "args": tc.get("arguments", {}),
                }
                for tc in turn.get("toolCalls", [])
            ]
        }
        turnDetails.append(turnDetail)
    
    turnDetailsJson = json.dumps(turnDetails, indent=2, ensure_ascii=False)
    
    # Calculate cost
    totalInputTokens = runLog.get("totalInputTokens", 0)
    totalOutputTokens = runLog.get("totalOutputTokens", 0)
    estimatedCost = estimateCost(totalInputTokens, totalOutputTokens, model)
    costStr = f"{estimatedCost:.6f}" if estimatedCost is not None else "N/A"
    
    row = (
        runId,
        question["id"],
        model,
        reasoning,
        datetime.now().isoformat(),
        runLog["totalUserMessages"],
        runLog["totalToolCalls"],
        runLog.get("totalInputTokens", 0),
        runLog.get("totalOutputTokens", 0),
        costStr,
        f"{elapsed:.2f}",
        perToolCounts["plotManagement"],
        perToolCounts["missionManagement"],
        perToolCounts["cameraSensors"],
        perToolCounts["executeFlightScript"],
        interactionSeq,
        turnDetailsJson,
        "",
    )
    ws.append(row)
    
    appendDetailedTurns(wb, runId, question["id"], model, reasoning, turns)
    
    wb.save(logPath)


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Pass^k Tool-Call Experiment Runner (Refactored)")
    parser.add_argument("--tools-base-url", default="http://localhost:3000", help="Base URL for Next.js tools API")
    parser.add_argument("--question-set-path", default=None, help="Path to question set Excel file")
    parser.add_argument("--log-path", default=None, help="Path to output Excel log file")
    parser.add_argument("--model", default=None, help="Model to use (default: run all configurations)")
    parser.add_argument("--reasoning", default=None, help="Reasoning effort: none, high (default: run all configurations)")
    parser.add_argument("--runs-per-config", type=int, default=1, help="Number of runs per configuration (default: 1)")
    parser.add_argument("--start-run", type=int, default=1, help="Start from run number (1-indexed, default: 1)")
    args = parser.parse_args()
    
    # Default paths
    scriptDir = os.path.dirname(os.path.abspath(__file__))
    questionSetPath = args.question_set_path or os.path.join(scriptDir, "data", "pass_k_question_set.xlsx")
    logPath = args.log_path or os.path.join(scriptDir, "pass_k_tool_call_log_v2.xlsx")
    
    if not os.path.isfile(questionSetPath):
        print(f"Question set not found: {questionSetPath}", file=sys.stderr)
        sys.exit(1)
    
    # Load questions
    print(f"Loading questions from: {questionSetPath}")
    questions = loadQuestionSet(questionSetPath)
    print(f"Loaded {len(questions)} questions.")
    
    if not questions:
        print("No questions to run.", file=sys.stderr)
        sys.exit(0)
    
    # Validate start-run
    if args.start_run < 1:
        print(f"Error: --start-run must be at least 1 (got {args.start_run})", file=sys.stderr)
        sys.exit(1)
    
    if args.start_run > args.runs_per_config:
        print(f"Warning: --start-run ({args.start_run}) > --runs-per-config ({args.runs_per_config})", file=sys.stderr)
        print(f"No runs will be executed. Adjust --start-run or increase --runs-per-config.", file=sys.stderr)
        sys.exit(1)
    
    # Define model configurations
    if args.model and args.reasoning:
        configurations = [(args.model, args.reasoning)]
    else:
        configurations = [
            ("gpt-5.2", "none"),
            ("gpt-5.2", "high"),
            ("gpt-5.4", "none"),
            ("gpt-5.4", "high"),
        ]
    
    # Create tools client
    toolsClient = createToolsClient(baseUrl=args.tools_base_url)
    
    # Run experiments
    print(f"Tools API URL: {args.tools_base_url}")
    print(f"Log path: {logPath}")
    print(f"Configurations: {len(configurations)}, runs per config: {args.runs_per_config}")
    if args.start_run > 1:
        print(f"Starting from run: {args.start_run}")
    print()
    
    totalRuns = 0
    skippedRuns = 0
    
    for configIdx, (model, reasoning) in enumerate(configurations):
        configName = f"{model}-reasoning-{reasoning}"
        print(f"=== Configuration {configIdx + 1}/{len(configurations)}: {configName} ===")
        
        # Create OpenAI client for this config
        openaiClient = createOpenAiClient(model=model)
        
        for runNum in range(args.runs_per_config):
            actualRunNum = runNum + 1
            
            # Skip runs before start_run
            if actualRunNum < args.start_run:
                skippedRuns += 1
                print(f"Skipping run {actualRunNum} (starting from run {args.start_run})")
                continue
            
            for qIdx, question in enumerate(questions):
                totalRuns += 1
                runId = f"{configName}-run{actualRunNum}-q{qIdx + 1}"
                qId = question["id"] or f"Q{qIdx + 1}"
                print(f"[{runId}] Running question {qId}...")
                
                startTime = time.perf_counter()
                runLog = runOneQuestion(question, openaiClient, toolsClient)
                elapsed = time.perf_counter() - startTime
                
                # Calculate cost
                totalInputTokens = runLog.get('totalInputTokens', 0)
                totalOutputTokens = runLog.get('totalOutputTokens', 0)
                estimatedCost = estimateCost(totalInputTokens, totalOutputTokens, model)
                
                print(f"  Model: {model}, Reasoning: {reasoning}")
                print(f"  User messages: {runLog['totalUserMessages']}, Tool calls: {runLog['totalToolCalls']}")
                print(f"  Tokens: in={totalInputTokens}, out={totalOutputTokens}")
                if estimatedCost is not None:
                    print(f"  Est. Cost: ${estimatedCost:.6f}")
                print(f"  Per-tool: {runLog['perToolCounts']}")
                print(f"  Interaction sequence: {formatInteractionSequence(runLog['turns'])}")
                
                # Show turn details
                for turn in runLog['turns']:
                    turnInputTokens = turn.get('inputTokens', 0)
                    turnOutputTokens = turn.get('outputTokens', 0)
                    turnCost = estimateCost(turnInputTokens, turnOutputTokens, model)
                    
                    print(f"    Turn {turn['turnIndex']}: USER")
                    userMsg = turn['userMessage']
                    print(f"      → Question: {userMsg[:80]}..." if len(userMsg) > 80 else f"      → Question: {userMsg}")
                    respText = turn['responseText']
                    print(f"      → Response: {respText[:80]}..." if len(respText) > 80 else f"      → Response: {respText}")
                    print(f"      → Tokens: in={turnInputTokens}, out={turnOutputTokens}", end="")
                    if turnCost is not None:
                        print(f" (${turnCost:.6f})")
                    else:
                        print()
                    toolCalls = turn.get('toolCalls', [])
                    if toolCalls:
                        print(f"      → Tool calls: {', '.join([tc.get('toolName', 'unknown') for tc in toolCalls])}")
                
                print(f"  Time: {elapsed:.2f}s")
                
                callTestScript(runLog)
                
                appendRunToExcel(logPath, runId, question, runLog, model, reasoning, elapsed)
                print()
        
        print()
    
    print(f"All runs completed. Total runs: {totalRuns}")
    if skippedRuns > 0:
        print(f"Skipped runs: {skippedRuns} (runs 1-{args.start_run - 1})")
    print(f"Log saved to: {logPath}")


if __name__ == "__main__":
    main()
