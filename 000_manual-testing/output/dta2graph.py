#!/usr/bin/env python3
"""
Waypoint-count distribution across all runs for Field 1 and Field 2.
Points colored green (pass) or red (fail) based on AND Result column.
Control (expected) waypoint count = 20 for both fields.
Y-axis clipped to 0–80; values outside shown as outlier markers at the edge.
"""

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path

FIELD1_PATH = Path(__file__).parent / "Field 1 - chatgpt-API-call-tester_log.xlsx"
FIELD2_PATH = Path(__file__).parent / "Field 2 - chatgpt-API-call-tester_log.xlsx"
OUT_DIR = Path(__file__).parent

CONTROL_WAYPOINTS = 20
Y_MAX = 80

PASS_COLOR = "#2ca02c"
FAIL_COLOR = "#d62728"


def loadField1():
    wb = openpyxl.load_workbook(FIELD1_PATH, data_only=True)
    ws = wb["Chat log"]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name, wps, andResult = row[0], row[12], row[14]
        if name is None or wps is None:
            continue
        if not isinstance(wps, (int, float)) or wps < 1:
            continue
        passed = bool(andResult) if andResult is not None else False
        records.append({"waypoints": int(wps), "passed": passed})
    return records


def loadField2():
    wb = openpyxl.load_workbook(FIELD2_PATH, data_only=True)
    ws = wb["Chat log"]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name, wps, andResult = row[0], row[15], row[16]
        if name is None or wps is None:
            continue
        if not isinstance(wps, (int, float)) or wps < 1:
            continue
        passed = bool(andResult) if andResult is not None else False
        records.append({"waypoints": int(wps), "passed": passed})
    return records


def plotWaypointDistribution(field1, field2):
    fig, ax = plt.subplots(figsize=(7, 7))

    fields = ["Field 1", "Field 2"]
    allData = [field1, field2]
    xPositions = [0, 1]

    for xPos, records in zip(xPositions, allData):
        for r in records:
            wps = r["waypoints"]
            color = PASS_COLOR if r["passed"] else FAIL_COLOR
            isOutlier = wps > Y_MAX

            yPlot = Y_MAX if isOutlier else wps
            marker = "^" if isOutlier else "o"
            size = 80 if isOutlier else 55

            jitter = np.random.default_rng(
                hash(f"{xPos}-{wps}-{r['passed']}") % 2**31
            ).uniform(-0.15, 0.15)

            ax.scatter(
                xPos + jitter,
                yPlot,
                c=color,
                marker=marker,
                s=size,
                alpha=0.8,
                edgecolors="white",
                linewidths=0.6,
                zorder=3,
            )

            if isOutlier:
                ax.annotate(
                    f"{wps}",
                    (xPos + jitter, yPlot),
                    textcoords="offset points",
                    xytext=(6, 4),
                    fontsize=7,
                    color=color,
                    fontweight="bold",
                )

    ax.axhline(
        y=CONTROL_WAYPOINTS,
        color="black",
        linestyle="--",
        linewidth=1.5,
        alpha=0.7,
        label=f"Control ({CONTROL_WAYPOINTS} waypoints)",
        zorder=2,
    )

    from matplotlib.lines import Line2D
    legendHandles = [
        Line2D([0], [0], marker="o", color="w", markerfacecolor=PASS_COLOR,
               markersize=9, label="Pass (AND Result = True)"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor=FAIL_COLOR,
               markersize=9, label="Fail (AND Result = False)"),
        Line2D([0], [0], marker="^", color="w", markerfacecolor="gray",
               markersize=9, label=f"Outlier (>{Y_MAX})"),
        Line2D([0], [0], linestyle="--", color="black", linewidth=1.5,
               label=f"Control ({CONTROL_WAYPOINTS})"),
    ]
    ax.legend(handles=legendHandles, fontsize=10, framealpha=0.9, loc="upper right")

    ax.set_xticks(xPositions)
    ax.set_xticklabels(fields, fontsize=13, fontweight="bold")
    ax.set_ylabel("Waypoint Count", fontsize=13, fontweight="bold")
    ax.set_title(
        "Waypoint Count — All Runs Combined",
        fontsize=14,
        fontweight="bold",
    )
    ax.set_ylim(-2, Y_MAX + 5)
    ax.set_xlim(-0.5, 1.5)
    ax.grid(axis="y", alpha=0.3, linestyle="--")
    ax.set_axisbelow(True)

    fig.tight_layout()
    return fig


def printSummary(field1, field2):
    for fieldName, records in [("Field 1", field1), ("Field 2", field2)]:
        wps = [r["waypoints"] for r in records]
        nPass = sum(1 for r in records if r["passed"])
        nFail = len(records) - nPass
        arr = np.array(wps)
        print(f"\n{fieldName}: n={len(records)} (pass={nPass}, fail={nFail}), "
              f"median={np.median(arr):.0f}, mean={np.mean(arr):.1f}, "
              f"min={arr.min()}, max={arr.max()}")
        print(f"  Values: {sorted(wps)}")
    print(f"\nControl: {CONTROL_WAYPOINTS} waypoints")


def main():
    field1 = loadField1()
    field2 = loadField2()

    printSummary(field1, field2)

    fig = plotWaypointDistribution(field1, field2)
    outPath = OUT_DIR / "waypoint_distribution.png"
    fig.savefig(outPath, dpi=150, bbox_inches="tight")
    print(f"\nPlot saved to: {outPath}")

    plt.show()


if __name__ == "__main__":
    main()
