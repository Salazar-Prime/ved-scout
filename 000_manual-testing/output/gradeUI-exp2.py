#!/usr/bin/env python3
"""
Tkinter grading UI for Field 2 experiment results.
Loads the "Chat log" sheet from the Excel workbook and lets you grade
each row on four criteria (Direction, Overlap, Coverage, Height) plus
enter a waypoint count. Grades persist back to the same xlsx file.
"""
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, "Field 2 - chatgpt-API-call-tester_log.xlsx")

CRITERIA = ["Direction", "Overlap", "Coverage", "Height"]
CRITERIA_DEFAULTS = {"Direction": 0, "Overlap": 0, "Coverage": 0, "Height": 1}

COLOR_PASS = "#22883e"
COLOR_FAIL = "#cc2222"
COLOR_UNGRADED = "#888888"
COLOR_PARTIAL = "#cc9a00"


def _findOrCreateColumns(ws, headerRow) -> Dict[str, int]:
    """Find column indices (1-based) for grading columns, creating them if needed."""
    headerCells = [str(c.value).strip() if c.value else "" for c in headerRow]
    colMap: Dict[str, int] = {}
    nextCol = len(headerCells) + 1

    targets = CRITERIA + ["Waypoints", "Comments"]
    for name in targets:
        found = False
        for i, h in enumerate(headerCells):
            if h.lower() == name.lower():
                colMap[name] = i + 1
                found = True
                break
        if not found:
            colMap[name] = nextCol
            ws.cell(row=1, column=nextCol, value=name)
            headerCells.append(name)
            nextCol += 1
    return colMap


def loadRows(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Load all rows from the Chat log sheet."""
    allRows = list(ws.iter_rows(values_only=True))
    if not allRows:
        return [], []
    rawHeaders = [str(c).strip() if c else "" for c in allRows[0]]

    dataRows: List[Dict[str, Any]] = []
    for rowIdx, row in enumerate(allRows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d: Dict[str, Any] = {"_excelRow": rowIdx}
        for i in range(len(row)):
            if i < len(rawHeaders) and rawHeaders[i]:
                d[rawHeaders[i]] = row[i]
        dataRows.append(d)
    return rawHeaders, dataRows


class GradeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Field 2 — Grading UI")
        self.geometry("1400x800")
        self.minsize(1000, 600)

        self.filePath: Optional[str] = None
        self.rawHeaders: List[str] = []
        self.rows: List[Dict[str, Any]] = []
        self.currentIdx: int = -1
        self.colMap: Dict[str, int] = {}

        self.criteriaVars: Dict[str, tk.IntVar] = {}
        self.criteriaLabels: Dict[str, tk.Label] = {}
        self.waypointVar = tk.StringVar(value="")
        self.commentsVar = tk.StringVar(value="")

        self._buildWidgets()

        if os.path.isfile(DEFAULT_XLSX):
            self.filePath = DEFAULT_XLSX
            self.fileVar.set(DEFAULT_XLSX)
            self._loadFile()

    def _buildWidgets(self):
        topFrame = ttk.Frame(self, padding=4)
        topFrame.pack(fill=tk.X)

        ttk.Label(topFrame, text="Excel:").pack(side=tk.LEFT)
        self.fileVar = tk.StringVar(value=DEFAULT_XLSX)
        ttk.Entry(topFrame, textvariable=self.fileVar, width=80).pack(side=tk.LEFT, padx=4)
        ttk.Button(topFrame, text="Open...", command=self._browseFile).pack(side=tk.LEFT)
        ttk.Button(topFrame, text="Reload", command=self._loadFile).pack(side=tk.LEFT, padx=4)

        paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # --- left panel: row list ---
        leftFrame = ttk.Frame(paned, width=320)
        paned.add(leftFrame, weight=0)

        ttk.Label(leftFrame, text="Rows", font=("TkDefaultFont", 11, "bold")).pack(anchor=tk.W)

        filterFrame = ttk.Frame(leftFrame)
        filterFrame.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(filterFrame, text="Filter:").pack(side=tk.LEFT)
        self.filterVar = tk.StringVar(value="All")
        filterCombo = ttk.Combobox(filterFrame, textvariable=self.filterVar, state="readonly", width=18)
        filterCombo["values"] = ["All", "Ungraded", "Partially Graded", "Fully Graded"]
        filterCombo.pack(side=tk.LEFT, padx=4)
        filterCombo.bind("<<ComboboxSelected>>", lambda _: self._populateRowList())

        listFrame = ttk.Frame(leftFrame)
        listFrame.pack(fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(listFrame, orient=tk.VERTICAL)
        self.rowListbox = tk.Listbox(listFrame, yscrollcommand=scrollbar.set, font=("TkFixedFont", 10))
        scrollbar.config(command=self.rowListbox.yview)
        self.rowListbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.rowListbox.bind("<<ListboxSelect>>", self._onRowSelect)

        # --- right panel: detail + grading ---
        rightFrame = ttk.Frame(paned)
        paned.add(rightFrame, weight=1)

        detailPaned = ttk.PanedWindow(rightFrame, orient=tk.VERTICAL)
        detailPaned.pack(fill=tk.BOTH, expand=True)

        # input pane
        inputFrame = ttk.LabelFrame(rightFrame, text="Input Prompt", padding=4)
        detailPaned.add(inputFrame, weight=1)
        self.inputText = tk.Text(inputFrame, wrap=tk.WORD, state=tk.DISABLED, height=8, font=("TkFixedFont", 10))
        inputScroll = ttk.Scrollbar(inputFrame, orient=tk.VERTICAL, command=self.inputText.yview)
        self.inputText.configure(yscrollcommand=inputScroll.set)
        self.inputText.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inputScroll.pack(side=tk.LEFT, fill=tk.Y)

        # output pane
        outputFrame = ttk.LabelFrame(rightFrame, text="Model Output", padding=4)
        detailPaned.add(outputFrame, weight=2)
        self.outputText = tk.Text(outputFrame, wrap=tk.WORD, state=tk.DISABLED, height=16, font=("TkFixedFont", 10))
        outputScroll = ttk.Scrollbar(outputFrame, orient=tk.VERTICAL, command=self.outputText.yview)
        self.outputText.configure(yscrollcommand=outputScroll.set)
        self.outputText.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        outputScroll.pack(side=tk.LEFT, fill=tk.Y)

        # metadata bar
        metaFrame = ttk.Frame(rightFrame, padding=(4, 2))
        metaFrame.pack(fill=tk.X)
        self.metaVar = tk.StringVar(value="")
        ttk.Label(metaFrame, textvariable=self.metaVar, font=("TkDefaultFont", 10)).pack(side=tk.LEFT)

        # --- grading controls ---
        gradeOuterFrame = ttk.LabelFrame(rightFrame, text="Grading", padding=8)
        gradeOuterFrame.pack(fill=tk.X, padx=4, pady=(4, 0))

        criteriaFrame = ttk.Frame(gradeOuterFrame)
        criteriaFrame.pack(fill=tk.X)

        for i, criterion in enumerate(CRITERIA):
            cFrame = ttk.Frame(criteriaFrame)
            cFrame.pack(side=tk.LEFT, padx=(0, 24))

            ttk.Label(cFrame, text=criterion, font=("TkDefaultFont", 12, "bold")).pack(anchor=tk.W)
            var = tk.IntVar(value=CRITERIA_DEFAULTS[criterion])
            self.criteriaVars[criterion] = var

            defVal = CRITERIA_DEFAULTS[criterion]
            statusLabel = tk.Label(
                cFrame,
                text="PASS" if defVal == 1 else "FAIL",
                width=8,
                bg=COLOR_PASS if defVal == 1 else COLOR_FAIL,
                fg="white",
                font=("TkDefaultFont", 13, "bold"),
                relief=tk.FLAT, padx=6, pady=4,
            )
            statusLabel.pack(anchor=tk.W, pady=(2, 2))
            self.criteriaLabels[criterion] = statusLabel

            btnFrame = ttk.Frame(cFrame)
            btnFrame.pack(anchor=tk.W)
            ttk.Button(
                btnFrame, text="Pass",
                command=lambda c=criterion: self._setCriterion(c, 1),
            ).pack(side=tk.LEFT, padx=(0, 4))
            ttk.Button(
                btnFrame, text="Fail",
                command=lambda c=criterion: self._setCriterion(c, 0),
            ).pack(side=tk.LEFT)

        # waypoints input
        wpFrame = ttk.Frame(criteriaFrame)
        wpFrame.pack(side=tk.LEFT, padx=(0, 24))
        ttk.Label(wpFrame, text="Waypoints", font=("TkDefaultFont", 12, "bold")).pack(anchor=tk.W)
        wpEntry = ttk.Entry(wpFrame, textvariable=self.waypointVar, width=8, font=("TkDefaultFont", 12))
        wpEntry.pack(anchor=tk.W)
        wpEntry.bind("<Return>", lambda _: self._saveAndNext())

        # comments input
        commentFrame = ttk.Frame(criteriaFrame)
        commentFrame.pack(side=tk.LEFT, padx=(0, 24), fill=tk.X, expand=True)
        ttk.Label(commentFrame, text="Comments", font=("TkDefaultFont", 12, "bold")).pack(anchor=tk.W)
        commentEntry = ttk.Entry(commentFrame, textvariable=self.commentsVar, width=30, font=("TkDefaultFont", 12))
        commentEntry.pack(anchor=tk.W, fill=tk.X, expand=True)
        commentEntry.bind("<Return>", lambda _: self._saveAndNext())

        # save + nav
        actionFrame = ttk.Frame(gradeOuterFrame)
        actionFrame.pack(fill=tk.X, pady=(8, 0))

        ttk.Button(actionFrame, text="< Prev (p)", command=self._prevRow).pack(side=tk.LEFT)
        ttk.Button(actionFrame, text="Next > (n)", command=self._nextRow).pack(side=tk.LEFT, padx=(4, 16))

        ttk.Button(
            actionFrame, text="Save & Next (Enter/s)",
            command=self._saveAndNext,
        ).pack(side=tk.LEFT, padx=4)

        self.statusVar = tk.StringVar(value="No file loaded")
        ttk.Label(actionFrame, textvariable=self.statusVar).pack(side=tk.RIGHT)

        # # keyboard shortcuts
        # self.bind("s", lambda _: self._saveAndNext())
        # self.bind("n", lambda _: self._nextRow())
        # self.bind("p", lambda _: self._prevRow())
        # self.bind("<Return>", lambda _: self._saveAndNext())

        # # criteria keyboard shortcuts: d=Direction, o=Overlap, c=Coverage, h=Height toggle
        # self.bind("d", lambda _: self._toggleCriterion("Direction"))
        # self.bind("o", lambda _: self._toggleCriterion("Overlap"))
        # self.bind("c", lambda _: self._toggleCriterion("Coverage"))
        # self.bind("h", lambda _: self._toggleCriterion("Height"))

    # ----- file ops -----

    def _browseFile(self):
        path = filedialog.askopenfilename(
            initialdir=SCRIPT_DIR,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.fileVar.set(path)
            self.filePath = path
            self._loadFile()

    def _loadFile(self):
        path = self.fileVar.get()
        if not path or not os.path.isfile(path):
            messagebox.showerror("Error", f"File not found:\n{path}")
            return
        self.filePath = path

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open workbook:\n{e}")
            return

        sheetName = None
        for name in wb.sheetnames:
            if name.lower() in ("chat log", "chatlog"):
                sheetName = name
                break
        if not sheetName:
            sheetName = wb.sheetnames[0]

        ws = wb[sheetName]

        self.colMap = _findOrCreateColumns(ws, ws[1])
        wb.save(path)

        self.rawHeaders, self.rows = loadRows(ws)
        wb.close()

        self._populateRowList()
        self.statusVar.set(f"Loaded {len(self.rows)} rows from '{sheetName}'  |  {os.path.basename(path)}")

    # ----- row grading status -----

    def _rowGradeStatus(self, row: Dict[str, Any]) -> str:
        """Return 'ungraded', 'partial', or 'full'."""
        criteriaFilled = 0
        for c in CRITERIA:
            val = row.get(c)
            if val is not None and str(val).strip() != "":
                criteriaFilled += 1
        if criteriaFilled == 0:
            return "ungraded"
        elif criteriaFilled == len(CRITERIA):
            return "full"
        else:
            return "partial"

    # ----- row list -----

    def _filteredIndices(self) -> List[int]:
        filt = self.filterVar.get()
        indices = []
        for i, r in enumerate(self.rows):
            status = self._rowGradeStatus(r)
            if filt == "All":
                indices.append(i)
            elif filt == "Ungraded" and status == "ungraded":
                indices.append(i)
            elif filt == "Partially Graded" and status == "partial":
                indices.append(i)
            elif filt == "Fully Graded" and status == "full":
                indices.append(i)
        return indices

    def _populateRowList(self):
        self.rowListbox.delete(0, tk.END)
        self._visibleIndices = self._filteredIndices()
        for listIdx, vi in enumerate(self._visibleIndices):
            r = self.rows[vi]
            name = r.get("Name", "?")
            status = self._rowGradeStatus(r)
            if status == "full":
                tag = " [graded]"
                color = COLOR_PASS
            elif status == "partial":
                tag = " [partial]"
                color = COLOR_PARTIAL
            else:
                tag = ""
                color = COLOR_UNGRADED
            label = f"#{vi+1}  {name}{tag}"
            self.rowListbox.insert(tk.END, label)
            self.rowListbox.itemconfig(listIdx, fg=color)

        if self._visibleIndices:
            self.rowListbox.selection_set(0)
            self.rowListbox.event_generate("<<ListboxSelect>>")
        else:
            self.currentIdx = -1
            self._clearDetail()

    # ----- row selection -----

    def _onRowSelect(self, _event=None):
        sel = self.rowListbox.curselection()
        if not sel:
            return
        listIdx = sel[0]
        if listIdx >= len(self._visibleIndices):
            return
        self.currentIdx = self._visibleIndices[listIdx]
        self._showDetail(self.currentIdx)

    def _prevRow(self):
        sel = self.rowListbox.curselection()
        if not sel or sel[0] == 0:
            return
        newSel = sel[0] - 1
        self.rowListbox.selection_clear(0, tk.END)
        self.rowListbox.selection_set(newSel)
        self.rowListbox.see(newSel)
        self.rowListbox.event_generate("<<ListboxSelect>>")

    def _nextRow(self):
        sel = self.rowListbox.curselection()
        if not sel or sel[0] >= self.rowListbox.size() - 1:
            return
        newSel = sel[0] + 1
        self.rowListbox.selection_clear(0, tk.END)
        self.rowListbox.selection_set(newSel)
        self.rowListbox.see(newSel)
        self.rowListbox.event_generate("<<ListboxSelect>>")

    # ----- detail display -----

    def _clearDetail(self):
        for w in (self.inputText, self.outputText):
            w.configure(state=tk.NORMAL)
            w.delete("1.0", tk.END)
            w.configure(state=tk.DISABLED)
        self.metaVar.set("")
        for c in CRITERIA:
            self.criteriaVars[c].set(CRITERIA_DEFAULTS[c])
        self.waypointVar.set("")
        self.commentsVar.set("")

    def _showDetail(self, idx: int):
        row = self.rows[idx]

        # input
        self.inputText.configure(state=tk.NORMAL)
        self.inputText.delete("1.0", tk.END)
        self.inputText.insert(tk.END, str(row.get("Input", "")))
        self.inputText.configure(state=tk.DISABLED)

        # output
        self.outputText.configure(state=tk.NORMAL)
        self.outputText.delete("1.0", tk.END)
        self.outputText.insert(tk.END, str(row.get("Output", "")))
        self.outputText.configure(state=tk.DISABLED)

        # metadata
        name = row.get("Name", "")
        tokens = f"in={row.get('Input tokens', '?')}, out={row.get('Output tokens', '?')}"
        cost = row.get("Est. cost ($)", "?")
        time = row.get("Time (s)", "?")
        self.metaVar.set(f"{name}  |  Tokens: {tokens}  |  Cost: ${cost}  |  Time: {time}s")

        # load existing grades into UI
        for c in CRITERIA:
            existing = row.get(c)
            if existing is not None and str(existing).strip() != "":
                self.criteriaVars[c].set(int(existing))
            else:
                self.criteriaVars[c].set(CRITERIA_DEFAULTS[c])

        wp = row.get("Waypoints")
        if wp is not None and str(wp).strip() != "":
            self.waypointVar.set(str(int(wp)) if isinstance(wp, (int, float)) else str(wp))
        else:
            self.waypointVar.set("")

        comment = row.get("Comments")
        if comment is not None and str(comment).strip() != "":
            self.commentsVar.set(str(comment))
        else:
            self.commentsVar.set("")

        self._updateCriteriaIndicators()

    def _updateCriteriaIndicators(self):
        for c in CRITERIA:
            val = self.criteriaVars[c].get()
            lbl = self.criteriaLabels[c]
            if val == 1:
                lbl.configure(text="PASS", bg=COLOR_PASS, fg="white")
            else:
                lbl.configure(text="FAIL", bg=COLOR_FAIL, fg="white")

    # ----- grading actions -----

    def _setCriterion(self, criterion: str, value: int):
        self.criteriaVars[criterion].set(value)
        self._updateCriteriaIndicators()

    def _toggleCriterion(self, criterion: str):
        if self.focus_get() and isinstance(self.focus_get(), (tk.Entry, ttk.Entry)):
            return
        current = self.criteriaVars[criterion].get()
        self.criteriaVars[criterion].set(0 if current == 1 else 1)
        self._updateCriteriaIndicators()

    def _saveAndNext(self):
        if self.currentIdx < 0 or not self.filePath:
            return

        if self.focus_get() and isinstance(self.focus_get(), (tk.Entry, ttk.Entry)):
            pass

        row = self.rows[self.currentIdx]
        excelRow = row["_excelRow"]

        wpText = self.waypointVar.get().strip()
        wpValue = None
        if wpText:
            try:
                wpValue = int(wpText)
            except ValueError:
                messagebox.showerror("Error", "Waypoints must be a number.")
                return

        commentsText = self.commentsVar.get().strip()

        gradeValues = {}
        for c in CRITERIA:
            gradeValues[c] = self.criteriaVars[c].get()
        gradeValues["Waypoints"] = wpValue
        if commentsText:
            gradeValues["Comments"] = commentsText

        try:
            self._saveToExcel(excelRow, gradeValues)
        except PermissionError:
            messagebox.showerror("Error", "Cannot save — file may be open in another app. Close it and retry.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Save failed:\n{e}")
            return

        for c in CRITERIA:
            row[c] = gradeValues[c]
        if wpValue is not None:
            row["Waypoints"] = wpValue
        if commentsText:
            row["Comments"] = commentsText

        sel = self.rowListbox.curselection()
        if sel:
            listIdx = sel[0]
            name = row.get("Name", "?")
            status = self._rowGradeStatus(row)
            if status == "full":
                tag = " [graded]"
                color = COLOR_PASS
            elif status == "partial":
                tag = " [partial]"
                color = COLOR_PARTIAL
            else:
                tag = ""
                color = COLOR_UNGRADED
            label = f"#{self.currentIdx+1}  {name}{tag}"
            self.rowListbox.delete(listIdx)
            self.rowListbox.insert(listIdx, label)
            self.rowListbox.itemconfig(listIdx, fg=color)
            self.rowListbox.selection_set(listIdx)

        passCount = sum(1 for c in CRITERIA if gradeValues[c] == 1)
        self.statusVar.set(
            f"Saved row #{self.currentIdx+1}: {passCount}/{len(CRITERIA)} pass"
            + (f", {wpValue} waypoints" if wpValue is not None else "")
            + "  — saved"
        )

        self.after(150, self._nextRow)

    def _saveToExcel(self, excelRow: int, values: Dict[str, Any]):
        wb = load_workbook(self.filePath)
        sheetName = None
        for name in wb.sheetnames:
            if name.lower() in ("chat log", "chatlog"):
                sheetName = name
                break
        if not sheetName:
            sheetName = wb.sheetnames[0]
        ws = wb[sheetName]

        self.colMap = _findOrCreateColumns(ws, ws[1])

        for key, val in values.items():
            if val is not None and key in self.colMap:
                ws.cell(row=excelRow, column=self.colMap[key], value=val)

        wb.save(self.filePath)
        wb.close()


if __name__ == "__main__":
    app = GradeApp()
    app.mainloop()
