#!/usr/bin/env python3
"""
Generate a single KML file from chatgpt-API-call-tester_log.xlsx.
- One Placemark per row, named from the Name column.
- Output column content is included in the placemark.
- All other columns (Timestamp, Input, Input tokens, Output tokens, Est. cost, Time)
  are stored in the placemark description as metadata.
- If Output contains waypoint-like coordinates, they are parsed as UAS flight waypoints,
  numbered 1 to n in list order, plus a path LineString.

Requires: pip install openpyxl
Usage: python generate-kml-files-4-experiments.py [--input path] [--output path]
"""

import argparse
import json
import os
import re
import xml.etree.ElementTree as ET

DEFAULT_ALTITUDE_M = 10

# Regex: two numbers (optionally with leading "n." or "- "), optional third for altitude
COORD_PAIR_RE = re.compile(
    r"(?:^|\s)(?:(\d+)[.)]\s*)?"
    r"(-?\d+\.?\d*)\s*[,]\s*(-?\d+\.?\d*)"
    r"(?:\s*[,]\s*(-?\d+\.?\d*))?(?=\s|$|[)\]])",
    re.MULTILINE,
)

LOG_HEADERS = (
    "Name",
    "Timestamp",
    "Input",
    "Output",
    "Input tokens",
    "Output tokens",
    "Est. cost ($)",
    "Time (s)",
)
DEFAULT_LOG = "chatgpt-API-call-tester_log.xlsx"
DEFAULT_OUT_KML = "experiments_log.kml"


def parseWaypointsFromOutput(outputText, defaultAltM=DEFAULT_ALTITUDE_M):
    """
    Parse waypoint coordinates from Output text. Returns list of (lng, lat, alt) in order.
    Primary format: one waypoint per line, "lat, lng, alt" (e.g. "40.472647, -86.993611, 10").
    Also supports: GeoJSON, or generic coordinate pairs in text.
    """
    if not outputText or not str(outputText).strip():
        return []
    text = str(outputText).strip()

    # 1) Line-based format: "lat, lng, alt" or "lat, lng" per line (order preserved)
    coords = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("//"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) >= 2:
            try:
                lat = float(parts[0])
                lng = float(parts[1])
                alt = float(parts[2]) if len(parts) >= 3 else defaultAltM
                if -180 <= lng <= 180 and -90 <= lat <= 90:
                    coords.append((lng, lat, alt))
            except (ValueError, IndexError):
                pass
    if coords:
        return coords

    # 2) Try GeoJSON
    try:
        data = json.loads(text)
        coords = []
        if isinstance(data, dict):
            if data.get("type") == "FeatureCollection":
                for f in data.get("features", []):
                    geom = f.get("geometry") if isinstance(f, dict) else None
                    if geom and geom.get("type") == "Point":
                        c = geom.get("coordinates")
                        if c and len(c) >= 2:
                            alt = float(c[2]) if len(c) > 2 else defaultAltM
                            coords.append((float(c[0]), float(c[1]), alt))
            elif "coordinates" in data:
                raw = data["coordinates"]
                if isinstance(raw, list) and raw and isinstance(raw[0], (int, float)):
                    if len(raw) >= 2:
                        alt = float(raw[2]) if len(raw) > 2 else defaultAltM
                        coords.append((float(raw[0]), float(raw[1]), alt))
                else:
                    for c in raw:
                        if c and len(c) >= 2:
                            alt = float(c[2]) if len(c) > 2 else defaultAltM
                            coords.append((float(c[0]), float(c[1]), alt))
        if coords:
            return coords
    except (json.JSONDecodeError, TypeError, ValueError):
        pass

    # 3) Regex: find coordinate pairs in order (support lat,lng or lng,lat)
    coords = []
    for m in COORD_PAIR_RE.finditer(text):
        a, b = float(m.group(2)), float(m.group(3))
        alt = float(m.group(4)) if m.group(4) else defaultAltM
        # Prefer (lat,lng) when a looks like lat [-90,90] and b like lng [-180,180]
        if -90 <= a <= 90 and -180 <= b <= 180:
            lng, lat = b, a
        else:
            lng, lat = a, b
        if -180 <= lng <= 180 and -90 <= lat <= 90:
            coords.append((lng, lat, alt))
    return coords


def escapeForXml(text):
    if text is None:
        return ""
    s = str(text)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def buildDescription(row, headers):
    """Build description HTML: metadata table + Output content (all escaped for XML)."""
    parts = ['<table border="1" cellpadding="4">']
    for i, h in enumerate(headers):
        if h == "Output":
            continue
        val = row[i] if i < len(row) else ""
        if val is None:
            val = ""
        valStr = escapeForXml(str(val))
        parts.append(f"<tr><th>{escapeForXml(h)}</th><td>{valStr}</td></tr>")
    parts.append("</table>")
    outputVal = ""
    try:
        outIdx = headers.index("Output")
        if outIdx < len(row) and row[outIdx] is not None:
            outputVal = str(row[outIdx])
    except ValueError:
        pass
    if outputVal.strip():
        parts.append("<h3>Output</h3><pre>")
        parts.append(escapeForXml(outputVal))
        parts.append("</pre>")
    return "".join(parts)


def createPlacemark(name, description, pointCoords=None):
    """Create a Placemark with optional Point geometry. pointCoords = (lng, lat, alt) or None."""
    placemark = ET.Element("Placemark")
    ET.SubElement(placemark, "name").text = escapeForXml(name) if name else "Unnamed"
    descEl = ET.SubElement(placemark, "description")
    descEl.text = description
    if pointCoords is not None:
        lng, lat, alt = pointCoords
        point = ET.SubElement(placemark, "Point")
        ET.SubElement(point, "coordinates").text = f"{lng},{lat},{alt}"
    return placemark


def createWaypointPlacemark(index, lng, lat, alt):
    """Create a Placemark for a single UAS waypoint, named 1 to n."""
    placemark = ET.Element("Placemark")
    ET.SubElement(placemark, "name").text = str(index)
    point = ET.SubElement(placemark, "Point")
    ET.SubElement(point, "coordinates").text = f"{lng},{lat},{alt}"
    return placemark


def createPathPlacemark(coords):
    """Create a Placemark with LineString through waypoints. coords = list of (lng, lat, alt)."""
    placemark = ET.Element("Placemark")
    ET.SubElement(placemark, "name").text = "Path"
    line = ET.SubElement(placemark, "LineString")
    coordStr = "\n".join(f"{lon},{lat},{alt}" for lon, lat, alt in coords)
    ET.SubElement(line, "coordinates").text = coordStr
    return placemark


def main():
    scriptDir = os.path.dirname(os.path.abspath(__file__))
    parser = argparse.ArgumentParser(
        description="Generate one KML from chatgpt-API-call-tester log Excel."
    )
    parser.add_argument(
        "--input",
        "-i",
        default=os.path.join(scriptDir, DEFAULT_LOG),
        help=f"Input Excel log (default: {DEFAULT_LOG})",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=os.path.join(scriptDir, DEFAULT_OUT_KML),
        help=f"Output KML path (default: {DEFAULT_OUT_KML})",
    )
    parser.add_argument(
        "--altitude",
        "-a",
        type=float,
        default=DEFAULT_ALTITUDE_M,
        help=f"Default waypoint altitude in meters when not in Output (default: {DEFAULT_ALTITUDE_M})",
    )
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        return 1

    if not os.path.isfile(args.input):
        print(f"Input file not found: {args.input}")
        return 1

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Excel file has no rows.")
        return 1

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    nameIdx = 0
    if "Name" in headers:
        nameIdx = headers.index("Name")
    else:
        headers = list(LOG_HEADERS)

    kml = ET.Element("kml", attrib={"xmlns": "http://www.opengis.net/kml/2.2"})
    document = ET.SubElement(kml, "Document")
    ET.SubElement(document, "name").text = "Experiments log"

    outIdx = headers.index("Output") if "Output" in headers else -1

    for rowIndex, row in enumerate(rows[1:], 1):
        row = list(row) if row else []
        name = row[nameIdx] if nameIdx < len(row) else ""
        if name is None:
            name = ""
        name = str(name).strip() or f"Row_{rowIndex}"
        description = buildDescription(row, headers)

        outputText = ""
        if outIdx >= 0 and outIdx < len(row) and row[outIdx] is not None:
            outputText = str(row[outIdx])
        waypoints = parseWaypointsFromOutput(outputText, defaultAltM=args.altitude)

        if waypoints:
            folder = ET.SubElement(document, "Folder")
            ET.SubElement(folder, "name").text = escapeForXml(name)
            for i, (lng, lat, alt) in enumerate(waypoints, 1):
                folder.append(createWaypointPlacemark(i, lng, lat, alt))
            folder.append(createPathPlacemark(waypoints))
            folder.append(createPlacemark(f"{name} (metadata)", description))
        else:
            document.append(createPlacemark(name, description))

    tree = ET.ElementTree(kml)
    ET.indent(tree, space="  ")
    tree.write(args.output, encoding="unicode", method="xml")
    print(f"KML saved to {args.output} ({len(rows) - 1} placemarks)")
    return 0


if __name__ == "__main__":
    exit(main())
